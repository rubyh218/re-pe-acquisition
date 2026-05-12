"""
excel_writer.py -- Institutional XLSX output for infrastructure deals.

Sheets:
  - Executive Summary (shared writer, inserted as first tab)
  - Assumptions
  - Generation Schedule (gross, degradation, curtailment, availability, net)
  - Revenue Streams (per-stream year-by-year detail)
  - Pro Forma (revenue + opex + capex + debt build)
  - Debt
  - Returns
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

import scripts  # noqa: F401
from excel_style import (
    apply_institutional_styles,
    set_sheet_defaults,
    write_header,
    write_input,
    write_label,
    write_section,
)

from ..waterfall_acq import WaterfallResult
from .models import AvailabilityStream, MerchantStream, PPAStream
from .pro_forma import InfraProForma


# ---------------------------------------------------------------------------
# Assumptions
# ---------------------------------------------------------------------------

def _ws_assumptions(wb: Workbook, pf: InfraProForma) -> None:
    deal = pf.deal
    ws = wb.create_sheet("Assumptions")
    set_sheet_defaults(ws, f"{deal.deal_name} - Assumptions")
    write_header(ws, 1, f"{deal.deal_name} - Assumptions", span_cols=8)
    gen = deal.property.generation

    r = 3
    write_section(ws, r, "Property"); r += 2
    rows = [
        ("Asset Class", deal.property.asset_class, "general"),
        ("Technology", gen.technology, "general"),
        ("Market / ISO", deal.property.market, "general"),
        ("Submarket", deal.property.submarket, "general"),
        ("COD Date", deal.property.cod_date, "date"),
        ("Year Built", deal.property.year_built, "general"),
        ("Nameplate MW (AC)", gen.nameplate_mw_ac, "general"),
        ("Capacity Factor", gen.capacity_factor, "pct1"),
        ("Degradation %/yr", gen.degradation_pct, "pct2"),
        ("Curtailment %", gen.curtailment_pct, "pct1"),
        ("Availability %", gen.availability_pct, "pct1"),
        ("Yr-1 Gross Generation (MWh)", gen.gross_annual_generation_mwh_yr1, "general"),
    ]
    if gen.nameplate_mw_dc:
        rows.insert(7, ("Nameplate MW (DC)", gen.nameplate_mw_dc, "general"))
    if gen.technology == "bess":
        rows.extend([
            ("BESS Duration (hrs)", gen.bess_duration_hrs, "multiple"),
            ("BESS Cycles / yr", gen.bess_cycles_per_year, "general"),
            ("BESS Round-Trip Eff", gen.bess_round_trip_eff, "pct1"),
        ])
    for label, val, fmt in rows:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Acquisition"); r += 2
    for label, val, fmt in [
        ("Purchase Price", deal.acquisition.purchase_price, "dollar"),
        ("Price / MW", deal.acquisition.purchase_price / gen.nameplate_mw_ac, "dollar"),
        ("Closing Costs %", deal.acquisition.closing_costs_pct, "pct2"),
        ("Initial CapEx", deal.acquisition.initial_capex, "dollar"),
        ("Day-One Reserves", deal.acquisition.day_one_reserves, "dollar"),
        ("Close Date", deal.acquisition.close_date, "date"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Tax Credits"); r += 2
    for label, val, fmt in [
        ("ITC %", deal.tax_credits.itc_pct, "pct1"),
        ("ITC Basis", deal.tax_credits.itc_basis, "dollar"),
        ("ITC Cash (Yr 1)", deal.tax_credits.itc_pct * deal.tax_credits.itc_basis, "dollar"),
        ("PTC $/MWh", deal.tax_credits.ptc_per_mwh, "dollar"),
        ("PTC Term (yrs)", deal.tax_credits.ptc_term_yrs, "general"),
        ("PTC Inflation", deal.tax_credits.ptc_inflation, "pct1"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "OpEx"); r += 2
    for label, val, fmt in [
        ("Fixed O&M / MW", deal.opex.fixed_om_per_mw_yr, "dollar"),
        ("Variable O&M / MWh", deal.opex.variable_om_per_mwh, "dollar"),
        ("Insurance / MW", deal.opex.insurance_per_mw_yr, "dollar"),
        ("Property Tax (annual)", deal.opex.property_tax, "dollar"),
        ("Land Lease (annual)", deal.opex.land_lease, "dollar"),
        ("Interconnection O&M", deal.opex.interconnection_om, "dollar"),
        ("Asset Mgmt % of Revenue", deal.opex.asset_mgmt_pct, "pct1"),
        ("O&M Growth", deal.opex.om_growth, "pct1"),
        ("Prop Tax Growth", deal.opex.property_tax_growth, "pct1"),
        ("Land Lease Growth", deal.opex.land_lease_growth, "pct1"),
        ("Insurance Growth", deal.opex.insurance_growth, "pct1"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Debt"); r += 2
    for label, val, fmt in [
        ("Rate", deal.debt.rate, "pct2"),
        ("Term (yrs)", deal.debt.term_yrs, "general"),
        ("Amort (yrs)", deal.debt.amort_yrs, "general"),
        ("IO Period (yrs)", deal.debt.io_period_yrs, "general"),
        ("Max LTV", deal.debt.max_ltv, "pct1"),
        ("Min DSCR", deal.debt.min_dscr, "multiple"),
        ("Min Debt Yield", deal.debt.min_debt_yield, "pct1"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Exit"); r += 2
    for label, val, fmt in [
        ("Hold (yrs)", deal.exit.hold_yrs, "general"),
        ("Exit Cap", deal.exit.exit_cap, "pct2"),
        ("Cost of Sale %", deal.exit.cost_of_sale_pct, "pct2"),
        ("Exit NOI Basis", deal.exit.exit_noi_basis, "general"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1


# ---------------------------------------------------------------------------
# Generation schedule
# ---------------------------------------------------------------------------

def _ws_generation(wb: Workbook, pf: InfraProForma) -> None:
    ws = wb.create_sheet("Generation")
    set_sheet_defaults(ws, "Generation Schedule")
    write_header(ws, 1, f"{pf.deal.deal_name} - Generation Schedule", span_cols=10)
    headers = ["Year", "Gross MWh", "Degradation", "Curtailed MWh",
               "Avail. Loss MWh", "Net MWh", "Realized CF"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=3, column=i, value=h).style = "label_bold"
    r = 4
    for g in pf.generation_schedule:
        ws.cell(row=r, column=2, value=f"Yr {g.year}")
        write_input(ws, (r, 3), g.gross_mwh, fmt="general")
        write_input(ws, (r, 4), g.degradation_factor, fmt="pct2")
        write_input(ws, (r, 5), g.curtailed_mwh, fmt="general")
        write_input(ws, (r, 6), g.availability_loss_mwh, fmt="general")
        write_input(ws, (r, 7), g.net_mwh, fmt="general")
        write_input(ws, (r, 8), g.cf_realized, fmt="pct2")
        r += 1


# ---------------------------------------------------------------------------
# Revenue streams (per-stream year-by-year)
# ---------------------------------------------------------------------------

def _ws_revenue_streams(wb: Workbook, pf: InfraProForma) -> None:
    ws = wb.create_sheet("Revenue Streams")
    set_sheet_defaults(ws, "Revenue Streams")
    write_header(ws, 1, f"{pf.deal.deal_name} - Revenue by Stream", span_cols=14)
    n = len(pf.years)
    ws.cell(row=3, column=2, value="Stream").style = "label_bold"
    ws.cell(row=3, column=3, value="Counterparty").style = "label_bold"
    ws.cell(row=3, column=4, value="Kind").style = "label_bold"
    for y in range(1, n + 1):
        ws.cell(row=3, column=4 + y, value=f"Yr {y}").style = "label_bold"
    r = 4

    # Stream roster (one row per stream, totals across hold)
    for label, years in pf.per_stream_years.items():
        first = years[0]
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=3, value=first.counterparty)
        ws.cell(row=r, column=4, value=first.kind.upper())
        for i, sy in enumerate(years):
            write_input(ws, (r, 5 + i), sy.revenue, fmt="dollar")
        r += 1

    r += 1
    write_section(ws, r, "Contracted vs. Merchant"); r += 2
    ws.cell(row=r, column=2, value="").style = "label_bold"
    for y in range(1, n + 1):
        ws.cell(row=r, column=4 + y, value=f"Yr {y}").style = "label_bold"
    r += 1
    write_label(ws, f"B{r}", "Contracted Revenue (PPA + Avail)")
    for i, cs in enumerate(pf.contracted_share_schedule):
        write_input(ws, (r, 5 + i), cs.contracted_revenue, fmt="dollar")
    r += 1
    write_label(ws, f"B{r}", "Merchant Revenue")
    for i, cs in enumerate(pf.contracted_share_schedule):
        write_input(ws, (r, 5 + i), cs.merchant_revenue, fmt="dollar")
    r += 1
    write_label(ws, f"B{r}", "Contracted Share", bold=True)
    for i, cs in enumerate(pf.contracted_share_schedule):
        write_input(ws, (r, 5 + i), cs.contracted_share, fmt="pct1")


# ---------------------------------------------------------------------------
# Pro forma
# ---------------------------------------------------------------------------

def _ws_pro_forma(wb: Workbook, pf: InfraProForma) -> None:
    ws = wb.create_sheet("Pro Forma")
    set_sheet_defaults(ws, "Pro Forma")
    write_header(ws, 1, f"{pf.deal.deal_name} - Pro Forma", span_cols=12)
    n = len(pf.years)
    ws.cell(row=3, column=2, value="").style = "label_bold"
    for y in range(1, n + 1):
        ws.cell(row=3, column=2 + y, value=f"Yr {y}").style = "label_bold"
    r = 4

    def row(label, getter, fmt="dollar", bold=False):
        nonlocal r
        write_label(ws, f"B{r}", label, bold=bold)
        for i, yl in enumerate(pf.years):
            write_input(ws, (r, 3 + i), getter(yl), fmt=fmt)
        r += 1

    write_section(ws, r, "Generation"); r += 2
    row("Net Generation (MWh)", lambda y: y.net_generation_mwh, fmt="general")
    r += 1

    write_section(ws, r, "Revenue"); r += 2
    row("PPA Revenue", lambda y: y.ppa_revenue)
    row("Availability Revenue", lambda y: y.availability_revenue)
    row("Merchant Revenue", lambda y: y.merchant_revenue)
    row("PTC Cash", lambda y: y.ptc_revenue)
    row("Gross Revenue", lambda y: y.gross_revenue, bold=True)
    r += 1

    write_section(ws, r, "OpEx"); r += 2
    row("Fixed O&M", lambda y: -y.fixed_om)
    row("Variable O&M", lambda y: -y.variable_om)
    row("Insurance", lambda y: -y.insurance)
    row("Property Tax", lambda y: -y.property_tax)
    row("Land Lease", lambda y: -y.land_lease)
    row("Interconnection O&M", lambda y: -y.interconnection_om)
    row("Asset Management", lambda y: -y.asset_mgmt_fee)
    row("Total OpEx", lambda y: -y.total_opex, bold=True)
    row("NOI", lambda y: y.noi, bold=True)
    r += 1

    write_section(ws, r, "CapEx"); r += 2
    row("Augmentation", lambda y: -y.augmentation)
    row("Recurring Reserve", lambda y: -y.recurring_reserve)
    row("ITC Cash (Yr 1)", lambda y: y.itc_cash)
    row("NCF Unlevered", lambda y: y.ncf_unlevered, bold=True)
    row("Debt Service", lambda y: -y.debt_service)
    row("NCF Levered", lambda y: y.ncf_levered, bold=True)


# ---------------------------------------------------------------------------
# Debt / Returns (mirror DC patterns)
# ---------------------------------------------------------------------------

def _ws_debt(wb: Workbook, pf: InfraProForma) -> None:
    ws = wb.create_sheet("Debt")
    set_sheet_defaults(ws, "Debt Sizing")
    write_header(ws, 1, f"{pf.deal.deal_name} - Debt Sizing & Amort", span_cols=8)
    r = 3
    write_section(ws, r, "Sizing (Yr 1 NOI)"); r += 2
    s = pf.sizing
    for label, val, fmt in [
        ("Loan Amount", s.loan_amount, "dollar"),
        ("Binding Constraint", s.binding, "general"),
        ("Implied LTV", s.implied_ltv, "pct1"),
        ("Implied DSCR", s.implied_dscr, "multiple"),
        ("Implied Debt Yield", s.implied_debt_yield, "pct2"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1
    r += 1; write_section(ws, r, "Amortization"); r += 2
    for i, h in enumerate(["Yr", "Beg.", "Interest", "Principal", "Debt Svc", "End", "IO?"], start=2):
        ws.cell(row=r, column=i, value=h).style = "label_bold"
    r += 1
    for am in pf.amort_schedule:
        ws.cell(row=r, column=2, value=am.year)
        write_input(ws, (r, 3), am.beginning_balance, fmt="dollar")
        write_input(ws, (r, 4), am.interest, fmt="dollar")
        write_input(ws, (r, 5), am.principal, fmt="dollar")
        write_input(ws, (r, 6), am.debt_service, fmt="dollar")
        write_input(ws, (r, 7), am.ending_balance, fmt="dollar")
        ws.cell(row=r, column=8, value="Yes" if am.is_io else "No")
        r += 1


def _ws_returns(wb: Workbook, pf: InfraProForma, wf: WaterfallResult) -> None:
    ws = wb.create_sheet("Returns")
    set_sheet_defaults(ws, "Returns")
    write_header(ws, 1, f"{pf.deal.deal_name} - Returns", span_cols=8)
    r = 3
    write_section(ws, r, "Sources & Uses"); r += 2
    su = pf.sources_uses
    for label, val in [
        ("Purchase Price", su.purchase_price),
        ("Closing Costs", su.closing_costs),
        ("Initial CapEx", su.initial_capex),
        ("Day-One Reserves", su.day_one_reserves),
        ("Acq Fee", su.acq_fee),
        ("Origination Fee", su.origination_fee),
        ("Lender Reserves", su.lender_reserves),
        ("Total Uses", su.total_uses),
        ("Loan Amount", su.loan_amount),
        ("Equity Check", su.equity_check),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt="dollar"); r += 1
    r += 1; write_section(ws, r, "Exit"); r += 2
    ex = pf.exit_summary
    for label, val, fmt in [
        ("Exit Year", ex.exit_year, "general"),
        ("Exit NOI", ex.exit_noi, "dollar"),
        ("Exit Cap", ex.exit_cap, "pct2"),
        ("Gross Sale", ex.gross_sale, "dollar"),
        ("Cost of Sale", ex.cost_of_sale, "dollar"),
        ("Loan Payoff", ex.loan_payoff, "dollar"),
        ("Net Proceeds", ex.net_proceeds, "dollar"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1
    r += 1; write_section(ws, r, "Returns"); r += 2
    for label, val, fmt in [
        ("Going-In Cap", pf.going_in_cap, "pct2"),
        ("Stabilized Cap", pf.stabilized_cap, "pct2"),
        ("Untrended ROC @ Stab", pf.roc.untrended_stab, "pct2"),
        ("Trended ROC @ Stab", pf.roc.trended_stab, "pct2"),
        ("Exit FTM ROC", pf.roc.exit_ftm, "pct2"),
        ("All-In / MW", pf.all_in_basis_per_mw, "dollar"),
        ("Project IRR", wf.total_equity_irr, "pct2"),
        ("Project MOIC", wf.total_equity_moic, "multiple"),
        ("LP IRR", wf.lp.irr, "pct2"),
        ("LP MOIC", wf.lp.moic, "multiple"),
        ("GP IRR", wf.gp.irr, "pct2"),
        ("GP MOIC", wf.gp.moic, "multiple"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1


# ---------------------------------------------------------------------------
# Top-level
# ---------------------------------------------------------------------------

def write_infrastructure_workbook(pf: InfraProForma, wf: WaterfallResult, out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    apply_institutional_styles(wb)
    _ws_assumptions(wb, pf)
    _ws_generation(wb, pf)
    _ws_revenue_streams(wb, pf)
    _ws_pro_forma(wb, pf)
    _ws_debt(wb, pf)
    _ws_returns(wb, pf, wf)
    from ..excel_summary import build_payload, write_executive_summary
    tech = pf.deal.property.generation.technology.title()
    payload = build_payload(
        pf=pf, wf=wf,
        asset_class=f"Infrastructure ({tech})",
        denom_label="Nameplate MW",
        denom_value=pf.deal.property.generation.nameplate_mw_ac,
        per_denom_label="/MW",
        per_denom_fmt="dollar",
        value_add_capex_total=sum(ev.amount for ev in pf.deal.capex.augmentation_schedule),
    )
    write_executive_summary(wb, payload)
    wb.save(out_path)
    return out_path
