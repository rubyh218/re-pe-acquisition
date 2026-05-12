"""
excel_writer.py -- Institutional XLSX output for data center deals (wholesale + colo).

Wholesale sheets:
  - Assumptions / Contract Roster / Per-Contract CFs / Pro Forma / Rollover / Debt / Returns
Colo sheets:
  - Assumptions / Cabinet Mix / Pro Forma / Debt / Returns

Both flavors share the Executive Summary tab written by excel_summary.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

import scripts  # noqa: F401
from excel_style import (
    apply_institutional_styles,
    set_sheet_defaults,
    write_header,
    write_input,
    write_label,
    write_section,
    write_subtotal,
)

from ..waterfall_acq import WaterfallResult
from .colo_pro_forma import ColoProForma
from .wholesale_pro_forma import WholesaleProForma


# ===========================================================================
# WHOLESALE
# ===========================================================================

def _ws_assumptions(wb: Workbook, pf: WholesaleProForma) -> None:
    deal = pf.deal
    ws = wb.create_sheet("Assumptions")
    set_sheet_defaults(ws, f"{deal.deal_name} - Assumptions")
    write_header(ws, 1, f"{deal.deal_name} - Assumptions", span_cols=8)

    r = 3
    write_section(ws, r, "Property"); r += 2
    for label, val, fmt in [
        ("Asset Class", deal.property.asset_class, "general"),
        ("Submarket", deal.property.submarket, "general"),
        ("Year Built", deal.property.year_built, "general"),
        ("Tier Rating", deal.property.tier_rating, "general"),
        ("Critical MW", deal.property.mw_critical, "general"),
        ("Commissioned MW", deal.property.mw_commissioned, "general"),
        ("Leased MW (in-place)", deal.property.leased_mw, "general"),
        ("Utilization", deal.property.utilization_pct, "pct1"),
        ("PUE", deal.property.pue, "multiple"),
        ("In-Place Annual Rent", deal.property.in_place_annual_rent, "dollar"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Acquisition"); r += 2
    for label, val, fmt in [
        ("Purchase Price", deal.acquisition.purchase_price, "dollar"),
        ("Price / MW", deal.acquisition.purchase_price / deal.property.mw_critical, "dollar"),
        ("Closing Costs %", deal.acquisition.closing_costs_pct, "pct2"),
        ("Initial CapEx (Day 1)", deal.acquisition.initial_capex, "dollar"),
        ("Day-One Reserves", deal.acquisition.day_one_reserves, "dollar"),
        ("Close Date", deal.acquisition.close_date, "date"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Market (Re-Leasing)"); r += 2
    for label, val, fmt in [
        ("Market Rent ($/kW/mo)", deal.market.market_rent_kw_mo, "dollar"),
        ("Market Rent Growth", deal.market.market_rent_growth, "pct1"),
        ("Renewal Probability", deal.market.renewal_prob, "pct1"),
        ("Downtime (mo, new tenant)", deal.market.downtime_mo, "general"),
        ("New TI / kW", deal.market.new_ti_kw, "dollar"),
        ("New LC %", deal.market.new_lc_pct, "pct1"),
        ("New Term (yrs)", deal.market.new_lease_term_yrs, "general"),
        ("New Escalation", deal.market.new_escalation_pct, "pct1"),
        ("Renewal Term (yrs)", deal.market.renewal_lease_term_yrs, "general"),
        ("Renewal Escalation", deal.market.renewal_escalation_pct, "pct1"),
        ("Power Margin Multiplier", deal.market.power_margin_multiplier, "multiple"),
        ("Utility Rate ($/kWh)", deal.market.utility_rate_kwh, "dollar"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "OpEx (sized on Critical MW)"); r += 2
    for label, val, fmt in [
        ("Security / MW", deal.opex.security_per_mw, "dollar"),
        ("MEP Staffing / MW", deal.opex.mep_staffing_per_mw, "dollar"),
        ("Insurance / MW", deal.opex.insurance_per_mw, "dollar"),
        ("Common Power / MW", deal.opex.common_power_per_mw, "dollar"),
        ("Non-Recoverable / MW", deal.opex.non_recoverable_per_mw, "dollar"),
        ("RE Tax (annual)", deal.opex.re_tax, "dollar"),
        ("Mgmt Fee % of EGI", deal.opex.mgmt_fee_pct, "pct1"),
        ("Controllable Growth", deal.opex.controllable_growth, "pct1"),
        ("RE Tax Growth", deal.opex.re_tax_growth, "pct1"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Debt"); r += 2
    for label, val, fmt in [
        ("Rate", deal.debt.rate, "pct2"),
        ("Term (yrs)", deal.debt.term_yrs, "general"),
        ("Amort (yrs)", deal.debt.amort_yrs, "general"),
        ("IO Period (yrs)", deal.debt.io_period_yrs, "general"),
        ("Max LTV", deal.debt.max_ltv, "pct1"),
        ("Min DSCR", deal.debt.min_dscr, "multiple"),
        ("Min Debt Yield", deal.debt.min_debt_yield, "pct1"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Exit"); r += 2
    for label, val, fmt in [
        ("Hold (yrs)", deal.exit.hold_yrs, "general"),
        ("Exit Cap", deal.exit.exit_cap, "pct2"),
        ("Cost of Sale %", deal.exit.cost_of_sale_pct, "pct2"),
        ("Exit NOI Basis", deal.exit.exit_noi_basis, "general"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1


def _ws_contract_roster(wb: Workbook, pf: WholesaleProForma) -> None:
    ws = wb.create_sheet("Contract Roster")
    set_sheet_defaults(ws, "Contract Roster")
    write_header(ws, 1, f"{pf.deal.deal_name} - Contract Roster", span_cols=10)
    headers = ["Tenant", "Suite", "MW", "$/kW/mo", "Annual Rent",
               "Lease Start", "Lease End", "Escalation", "Pass-Thru"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=3, column=i, value=h).style = "label_bold"
    r = 4
    for c in pf.deal.property.contracts:
        ws.cell(row=r, column=2, value=c.tenant)
        ws.cell(row=r, column=3, value=c.suite or "")
        write_input(ws, (r, 4), c.mw_leased, fmt="general")
        write_input(ws, (r, 5), c.base_rent_kw_mo, fmt="dollar")
        write_input(ws, (r, 6), c.annual_base_rent, fmt="dollar")
        write_input(ws, (r, 7), c.lease_start, fmt="date")
        write_input(ws, (r, 8), c.lease_end, fmt="date")
        write_input(ws, (r, 9), c.escalation_pct, fmt="pct1")
        ws.cell(row=r, column=10, value=c.power_pass_through)
        r += 1


def _ws_per_contract(wb: Workbook, pf: WholesaleProForma) -> None:
    ws = wb.create_sheet("Per-Contract CFs")
    set_sheet_defaults(ws, "Per-Contract CFs")
    write_header(ws, 1, f"{pf.deal.deal_name} - Per-Contract Cash Flows", span_cols=10)
    n = len(pf.years)
    ws.cell(row=3, column=2, value="Tenant").style = "label_bold"
    ws.cell(row=3, column=3, value="Line").style = "label_bold"
    for y in range(1, n + 1):
        c = ws.cell(row=3, column=3 + y, value=f"Yr {y}")
        c.style = "label_bold"
    r = 4
    for tenant_key, years in pf.per_contract_years.items():
        write_label(ws, f"B{r}", tenant_key, bold=True); r += 1
        for name, getter in [
            ("Base Rent", lambda cy: cy.base_rent),
            ("Free Rent", lambda cy: cy.free_rent),
            ("Power Margin", lambda cy: cy.power_margin),
            ("TI", lambda cy: -cy.ti),
            ("LC", lambda cy: -cy.lc),
            ("Occ. Mo.", lambda cy: cy.occupied_months),
        ]:
            ws.cell(row=r, column=3, value=name)
            for i, cy in enumerate(years):
                fmt = "general" if name == "Occ. Mo." else "dollar"
                write_input(ws, (r, 4 + i), getter(cy), fmt=fmt)
            r += 1
        r += 1


def _ws_wholesale_pf(wb: Workbook, pf: WholesaleProForma) -> None:
    ws = wb.create_sheet("Pro Forma")
    set_sheet_defaults(ws, "Pro Forma")
    write_header(ws, 1, f"{pf.deal.deal_name} - Pro Forma", span_cols=10)
    n = len(pf.years)
    ws.cell(row=3, column=2, value="").style = "label_bold"
    for y in range(1, n + 1):
        c = ws.cell(row=3, column=2 + y, value=f"Yr {y}")
        c.style = "label_bold"
    r = 4
    def row(label, getter, fmt="dollar", bold=False):
        nonlocal r
        write_label(ws, f"B{r}", label, bold=bold)
        for i, yl in enumerate(pf.years):
            write_input(ws, (r, 3 + i), getter(yl), fmt=fmt)
        r += 1
    write_section(ws, r, "Revenue"); r += 2
    row("Gross Rent", lambda y: y.gross_rent)
    row("Free Rent", lambda y: y.free_rent)
    row("Power Margin", lambda y: y.power_margin)
    row("General Vacancy", lambda y: y.general_vacancy)
    row("EGI", lambda y: y.egi, bold=True)
    r += 1
    write_section(ws, r, "OpEx"); r += 2
    row("Security", lambda y: -y.security)
    row("MEP Staffing", lambda y: -y.mep_staffing)
    row("Insurance", lambda y: -y.insurance)
    row("Common Power", lambda y: -y.common_power)
    row("RE Tax", lambda y: -y.re_tax)
    row("Non-Recoverable", lambda y: -y.non_recoverable)
    row("Mgmt Fee", lambda y: -y.mgmt_fee)
    row("Total OpEx", lambda y: -y.total_opex, bold=True)
    r += 1
    row("NOI", lambda y: y.noi, bold=True)
    r += 1
    write_section(ws, r, "CapEx"); r += 2
    row("TI", lambda y: -y.ti)
    row("LC", lambda y: -y.lc)
    row("Building CapEx + Reserves", lambda y: -y.building_capex)
    row("NCF Unlevered", lambda y: y.ncf_unlevered, bold=True)
    row("Debt Service", lambda y: -y.debt_service)
    row("NCF Levered", lambda y: y.ncf_levered, bold=True)
    r += 1
    row("Leased MW (avg)", lambda y: y.leased_mw_avg, fmt="general")
    row("Utilization", lambda y: y.utilization_pct, fmt="pct1")


def _ws_rollover(wb: Workbook, pf: WholesaleProForma) -> None:
    ws = wb.create_sheet("Rollover")
    set_sheet_defaults(ws, "Rollover Schedule")
    write_header(ws, 1, f"{pf.deal.deal_name} - Rollover Schedule", span_cols=8)
    headers = ["Year", "MW Rolling", "In-Place Rent", "Market Rent at Roll", "MTM Spread"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=3, column=i, value=h).style = "label_bold"
    r = 4
    for ro in pf.rollover_schedule:
        ws.cell(row=r, column=2, value=f"Yr {ro.year}")
        write_input(ws, (r, 3), ro.mw_rolling, fmt="general")
        write_input(ws, (r, 4), ro.in_place_rent_rolling, fmt="dollar")
        write_input(ws, (r, 5), ro.market_rent_at_roll, fmt="dollar")
        write_input(ws, (r, 6), ro.mtm_spread_pct, fmt="pct1")
        r += 1


# ===========================================================================
# COLO
# ===========================================================================

def _ws_colo_assumptions(wb: Workbook, pf: ColoProForma) -> None:
    deal = pf.deal
    ws = wb.create_sheet("Assumptions")
    set_sheet_defaults(ws, f"{deal.deal_name} - Assumptions")
    write_header(ws, 1, f"{deal.deal_name} - Assumptions", span_cols=8)
    r = 3
    write_section(ws, r, "Property"); r += 2
    for label, val, fmt in [
        ("Asset Class", deal.property.asset_class, "general"),
        ("Submarket", deal.property.submarket, "general"),
        ("Year Built", deal.property.year_built, "general"),
        ("Tier Rating", deal.property.tier_rating, "general"),
        ("Critical MW", deal.property.mw_critical, "general"),
        ("PUE", deal.property.pue, "multiple"),
        ("Total Cabinets", deal.property.total_cabinets, "general"),
        ("Total Contracted kW", deal.property.total_contracted_kw, "general"),
        ("In-Place Gross Rent", deal.property.in_place_gross_rent, "dollar"),
        ("Market Gross Rent", deal.property.market_gross_rent, "dollar"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1
    r += 1; write_section(ws, r, "Acquisition"); r += 2
    for label, val, fmt in [
        ("Purchase Price", deal.acquisition.purchase_price, "dollar"),
        ("Price / Cabinet", deal.acquisition.purchase_price / deal.property.total_cabinets, "dollar"),
        ("Price / MW", deal.acquisition.purchase_price / deal.property.mw_critical, "dollar"),
        ("Closing Costs %", deal.acquisition.closing_costs_pct, "pct2"),
        ("Initial CapEx", deal.acquisition.initial_capex, "dollar"),
        ("Day-One Reserves", deal.acquisition.day_one_reserves, "dollar"),
        ("Close Date", deal.acquisition.close_date, "date"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1
    r += 1; write_section(ws, r, "Revenue"); r += 2
    for label, val, fmt in [
        ("MRR Growth", deal.revenue.mrr_growth, "pct1"),
        ("Bad Debt %", deal.revenue.bad_debt, "pct1"),
        ("Yr 1 Concessions %", deal.revenue.concessions_yr1, "pct1"),
        ("XC per Cabinet", deal.revenue.xc_per_cabinet, "multiple"),
        ("XC MRR ($/mo)", deal.revenue.xc_mrr_each, "dollar"),
        ("Other Income / Cab / Mo", deal.revenue.other_income_per_cabinet_mo, "dollar"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1
    write_label(ws, f"B{r}", "Occupancy by Year"); r += 1
    for i, occ in enumerate(deal.revenue.occupancy, start=1):
        write_label(ws, f"B{r}", f"  Yr {i}"); write_input(ws, f"D{r}", occ, fmt="pct1"); r += 1
    r += 1; write_section(ws, r, "OpEx"); r += 2
    for label, val, fmt in [
        ("Utility Rate ($/kWh)", deal.opex.utility_rate_kwh, "dollar"),
        ("PUE Uplift (OpEx)", deal.opex.pue_uplift, "multiple"),
        ("Payroll / Cabinet", deal.opex.payroll_per_cabinet, "dollar"),
        ("R&M / Cabinet", deal.opex.rm_per_cabinet, "dollar"),
        ("Marketing / Cabinet", deal.opex.marketing_per_cabinet, "dollar"),
        ("Insurance / Cabinet", deal.opex.insurance_per_cabinet, "dollar"),
        ("Other / Cabinet", deal.opex.other_per_cabinet, "dollar"),
        ("RE Tax", deal.opex.re_tax, "dollar"),
        ("Mgmt Fee % of EGI", deal.opex.mgmt_fee_pct, "pct1"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1


def _ws_cabinet_mix(wb: Workbook, pf: ColoProForma) -> None:
    ws = wb.create_sheet("Cabinet Mix")
    set_sheet_defaults(ws, "Cabinet Mix")
    write_header(ws, 1, f"{pf.deal.deal_name} - Cabinet Mix", span_cols=8)
    headers = ["Cabinet Type", "Count", "kW / Cab", "Total kW",
               "In-Place MRR", "Market MRR", "Annual In-Place"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=3, column=i, value=h).style = "label_bold"
    r = 4
    for ct in pf.deal.property.cabinet_mix:
        ws.cell(row=r, column=2, value=ct.name)
        write_input(ws, (r, 3), ct.count, fmt="general")
        write_input(ws, (r, 4), ct.kw_per_cabinet, fmt="multiple")
        write_input(ws, (r, 5), ct.total_kw, fmt="general")
        write_input(ws, (r, 6), ct.in_place_mrr, fmt="dollar")
        write_input(ws, (r, 7), ct.market_mrr, fmt="dollar")
        write_input(ws, (r, 8), ct.in_place_annual_rent, fmt="dollar")
        r += 1


def _ws_colo_pf(wb: Workbook, pf: ColoProForma) -> None:
    ws = wb.create_sheet("Pro Forma")
    set_sheet_defaults(ws, "Pro Forma")
    write_header(ws, 1, f"{pf.deal.deal_name} - Pro Forma", span_cols=10)
    n = len(pf.years)
    ws.cell(row=3, column=2, value="").style = "label_bold"
    for y in range(1, n + 1):
        c = ws.cell(row=3, column=2 + y, value=f"Yr {y}")
        c.style = "label_bold"
    r = 4
    def row(label, getter, fmt="dollar", bold=False):
        nonlocal r
        write_label(ws, f"B{r}", label, bold=bold)
        for i, yl in enumerate(pf.years):
            write_input(ws, (r, 3 + i), getter(yl), fmt=fmt)
        r += 1
    write_section(ws, r, "Stats"); r += 2
    row("Occupancy", lambda y: y.occupancy, fmt="pct1")
    row("Occupied Cabinets", lambda y: y.occupied_cabinets, fmt="general")
    row("Avg MRR / Cabinet", lambda y: y.avg_mrr_per_cabinet)
    r += 1
    write_section(ws, r, "Revenue"); r += 2
    row("Cabinet Rent", lambda y: y.cabinet_rent)
    row("Cross-Connect Revenue", lambda y: y.cross_connect_rev)
    row("Other Income", lambda y: y.other_income)
    row("Gross Revenue", lambda y: y.gross_revenue, bold=True)
    row("Concessions", lambda y: y.concessions)
    row("Bad Debt", lambda y: y.bad_debt)
    row("EGI", lambda y: y.egi, bold=True)
    r += 1
    write_section(ws, r, "OpEx"); r += 2
    row("Power", lambda y: -y.power_cost)
    row("Payroll", lambda y: -y.payroll)
    row("R&M", lambda y: -y.rm)
    row("Marketing", lambda y: -y.marketing)
    row("Insurance", lambda y: -y.insurance)
    row("Other OpEx", lambda y: -y.other_opex)
    row("RE Tax", lambda y: -y.re_tax)
    row("Mgmt Fee", lambda y: -y.mgmt_fee)
    row("Total OpEx", lambda y: -y.total_opex, bold=True)
    row("NOI", lambda y: y.noi, bold=True)
    r += 1
    write_section(ws, r, "CapEx"); r += 2
    row("Fit-Out CapEx", lambda y: -y.fit_out_capex)
    row("Common CapEx", lambda y: -y.common_capex)
    row("Recurring Reserve", lambda y: -y.recurring_reserve)
    row("NCF Unlevered", lambda y: y.ncf_unlevered, bold=True)
    row("Debt Service", lambda y: -y.debt_service)
    row("NCF Levered", lambda y: y.ncf_levered, bold=True)


# ===========================================================================
# Shared (Debt / Returns)
# ===========================================================================

def _ws_debt(wb: Workbook, pf) -> None:
    ws = wb.create_sheet("Debt")
    set_sheet_defaults(ws, "Debt Sizing")
    write_header(ws, 1, f"{pf.deal.deal_name} - Debt Sizing & Amort", span_cols=8)
    r = 3
    write_section(ws, r, "Sizing (Yr 1 NOI)"); r += 2
    s = pf.sizing
    for label, val, fmt in [
        ("Loan Amount", s.loan_amount, "dollar"),
        ("Binding Constraint", s.binding, "general"),
        ("Implied LTV", s.implied_ltv, "pct1"),
        ("Implied DSCR", s.implied_dscr, "multiple"),
        ("Implied Debt Yield", s.implied_debt_yield, "pct2"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1
    r += 1; write_section(ws, r, "Amortization"); r += 2
    for i, h in enumerate(["Yr", "Beg.", "Interest", "Principal", "Debt Svc", "End", "IO?"], start=2):
        ws.cell(row=r, column=i, value=h).style = "label_bold"
    r += 1
    for am in pf.amort_schedule:
        ws.cell(row=r, column=2, value=am.year)
        write_input(ws, (r, 3), am.beginning_balance, fmt="dollar")
        write_input(ws, (r, 4), am.interest, fmt="dollar")
        write_input(ws, (r, 5), am.principal, fmt="dollar")
        write_input(ws, (r, 6), am.debt_service, fmt="dollar")
        write_input(ws, (r, 7), am.ending_balance, fmt="dollar")
        ws.cell(row=r, column=8, value="Yes" if am.is_io else "No")
        r += 1


def _ws_returns(wb: Workbook, pf, wf: WaterfallResult) -> None:
    ws = wb.create_sheet("Returns")
    set_sheet_defaults(ws, "Returns")
    write_header(ws, 1, f"{pf.deal.deal_name} - Returns", span_cols=8)
    r = 3
    write_section(ws, r, "Sources & Uses"); r += 2
    su = pf.sources_uses
    for label, val in [
        ("Purchase Price", su.purchase_price),
        ("Closing Costs", su.closing_costs),
        ("Initial CapEx", su.initial_capex),
        ("Day-One Reserves", su.day_one_reserves),
        ("Acq Fee", su.acq_fee),
        ("Origination Fee", su.origination_fee),
        ("Lender Reserves", su.lender_reserves),
        ("Total Uses", su.total_uses),
        ("Loan Amount", su.loan_amount),
        ("Equity Check", su.equity_check),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt="dollar"); r += 1
    r += 1; write_section(ws, r, "Exit"); r += 2
    ex = pf.exit_summary
    for label, val, fmt in [
        ("Exit Year", ex.exit_year, "general"),
        ("Exit NOI", ex.exit_noi, "dollar"),
        ("Exit Cap", ex.exit_cap, "pct2"),
        ("Gross Sale", ex.gross_sale, "dollar"),
        ("Cost of Sale", ex.cost_of_sale, "dollar"),
        ("Loan Payoff", ex.loan_payoff, "dollar"),
        ("Net Proceeds", ex.net_proceeds, "dollar"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1
    r += 1; write_section(ws, r, "Returns"); r += 2
    for label, val, fmt in [
        ("Going-In Cap", pf.going_in_cap, "pct2"),
        ("Stabilized Cap", pf.stabilized_cap, "pct2"),
        ("Project IRR", wf.total_equity_irr, "pct2"),
        ("Project MOIC", wf.total_equity_moic, "multiple"),
        ("LP IRR", wf.lp.irr, "pct2"),
        ("LP MOIC", wf.lp.moic, "multiple"),
        ("GP IRR", wf.gp.irr, "pct2"),
        ("GP MOIC", wf.gp.moic, "multiple"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1


# ===========================================================================
# Top-level
# ===========================================================================

def write_wholesale_workbook(pf: WholesaleProForma, wf: WaterfallResult, out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    apply_institutional_styles(wb)
    _ws_assumptions(wb, pf)
    _ws_contract_roster(wb, pf)
    _ws_per_contract(wb, pf)
    _ws_wholesale_pf(wb, pf)
    _ws_rollover(wb, pf)
    _ws_debt(wb, pf)
    _ws_returns(wb, pf, wf)
    from ..excel_summary import build_payload, write_executive_summary
    payload = build_payload(
        pf=pf, wf=wf,
        asset_class="Data Center (Wholesale)",
        denom_label="Critical MW",
        denom_value=pf.deal.property.mw_critical,
        per_denom_label="/MW",
        per_denom_fmt="dollar",
        value_add_capex_total=0.0,
    )
    write_executive_summary(wb, payload)
    wb.save(out_path)
    return out_path


def write_colo_workbook(pf: ColoProForma, wf: WaterfallResult, out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    apply_institutional_styles(wb)
    _ws_colo_assumptions(wb, pf)
    _ws_cabinet_mix(wb, pf)
    _ws_colo_pf(wb, pf)
    _ws_debt(wb, pf)
    _ws_returns(wb, pf, wf)
    from ..excel_summary import build_payload, write_executive_summary
    fit_out_total = (
        pf.deal.capex.fit_out_per_cabinet * pf.deal.property.total_cabinets
        if pf.deal.capex.fit_out_per_cabinet > 0 else 0.0
    )
    payload = build_payload(
        pf=pf, wf=wf,
        asset_class="Data Center (Colocation)",
        denom_label="Cabinets",
        denom_value=pf.deal.property.total_cabinets,
        per_denom_label="/Cabinet",
        per_denom_fmt="dollar",
        value_add_capex_total=fit_out_total,
    )
    write_executive_summary(wb, payload)
    wb.save(out_path)
    return out_path
