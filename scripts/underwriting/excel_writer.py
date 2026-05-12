"""
excel_writer.py — Institutional XLSX output with live formulas.

Sheets:
  - Assumptions     (all inputs, blue/yellow input cells)
  - Sources & Uses
  - Pro Forma       (Yr 1..hold[+1], with SUM formulas for EGI / OpEx / NOI / NCF)
  - Debt            (sizing analysis + annual amort schedule)
  - Returns         (XIRR/MOIC formulas referencing equity-flow row)
  - Sensitivities   (computed tables — static values referenced by labels)

Design: per-line per-year base assumptions are written as INPUTS (blue, yellow
fill). Subtotals, NOI, NCF, debt service totals, and IRR/MOIC are FORMULAS so
analysts can override any input cell and Excel recalculates downstream.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import scripts  # noqa: F401
from excel_style import (
    apply_institutional_styles,
    set_sheet_defaults,
    write_formula,
    write_header,
    write_input,
    write_label,
    write_link_internal,
    write_note,
    write_section,
    write_subheader,
    write_subtotal,
    write_total,
    write_units,
)

from .pro_forma import ProForma
from .sensitivity import SensitivityTable
from .waterfall_acq import WaterfallResult


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _col(idx: int) -> str:
    """1-indexed column number → letter (1=A, 4=D, ...)"""
    return get_column_letter(idx)


def _set_year_columns(ws, start_col: int, n_years: int, year_labels: list[str]):
    """Write year headers in row 4 starting at start_col."""
    for i, label in enumerate(year_labels):
        c = ws.cell(row=4, column=start_col + i, value=label)
        c.style = "label_bold"
        c.alignment = c.alignment.copy(horizontal="right")


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _write_assumptions(wb: Workbook, pf: ProForma):
    deal = pf.deal
    ws = wb.create_sheet("Assumptions")
    set_sheet_defaults(ws, f"{deal.deal_name} — Assumptions")
    write_header(ws, 1, f"{deal.deal_name} — Assumptions", span_cols=8)

    r = 3
    write_section(ws, r, "Property"); r += 2
    for label, val, fmt in [
        ("Asset Class", deal.property.asset_class, "general"),
        ("Submarket", deal.property.submarket, "general"),
        ("Year Built", deal.property.year_built, "general"),
        ("Unit Count", deal.property.unit_count, "general"),
        ("Total SF", deal.property.total_sf, "general"),
        ("In-Place GPR ($/yr)", deal.property.gpr_in_place, "dollar"),
        ("Market GPR ($/yr)", deal.property.gpr_market, "dollar"),
    ]:
        write_label(ws, f"B{r}", label)
        write_input(ws, f"D{r}", val, fmt=fmt)
        r += 1

    r += 1
    write_section(ws, r, "Acquisition"); r += 2
    for label, val, fmt in [
        ("Purchase Price", deal.acquisition.purchase_price, "dollar"),
        ("Closing Costs %", deal.acquisition.closing_costs_pct, "pct2"),
        ("Initial CapEx (Day 1)", deal.acquisition.initial_capex, "dollar"),
        ("Day-One Reserves", deal.acquisition.day_one_reserves, "dollar"),
        ("Close Date", deal.acquisition.close_date, "date"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1
    write_section(ws, r, "Revenue"); r += 2
    for label, val, fmt in [
        ("Other Income / Unit / Mo", deal.revenue.other_income_per_unit_mo, "dollar"),
        ("Other Income Growth", deal.revenue.other_income_growth, "pct1"),
        ("Vacancy", deal.revenue.vacancy, "pct1"),
        ("Bad Debt", deal.revenue.bad_debt, "pct1"),
        ("Concessions Yr 1", deal.revenue.concessions_yr1, "pct1"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1
    write_label(ws, f"B{r}", "Rent Growth (Yr 1..N)")
    for i, g in enumerate(deal.revenue.rent_growth):
        write_input(ws, (r, 4 + i), g, fmt="pct1")
    r += 2

    write_section(ws, r, "OpEx (per unit, except RE Tax)"); r += 2
    for label, val, fmt in [
        ("Payroll / unit", deal.opex.payroll_per_unit, "dollar"),
        ("R&M / unit", deal.opex.rm_per_unit, "dollar"),
        ("Marketing / unit", deal.opex.marketing_per_unit, "dollar"),
        ("Utilities / unit", deal.opex.utilities_per_unit, "dollar"),
        ("Insurance / unit", deal.opex.insurance_per_unit, "dollar"),
        ("Other / unit", deal.opex.other_per_unit, "dollar"),
        ("RE Tax (annual)", deal.opex.re_tax, "dollar"),
        ("RE Tax Growth", deal.opex.re_tax_growth, "pct1"),
        ("OpEx Growth (general)", deal.opex.growth, "pct1"),
        ("Mgmt Fee % of EGI", deal.opex.mgmt_fee_pct, "pct1"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1
    write_section(ws, r, "CapEx"); r += 2
    for label, val, fmt in [
        ("Value-Add / Unit", deal.capex.value_add_per_unit, "dollar"),
        ("Rent Premium / Unit / Mo", deal.capex.rent_premium_per_unit_mo, "dollar"),
        ("Common Area CapEx (Yr 1)", deal.capex.common_area_capex, "dollar"),
        ("Recurring Reserve / Unit", deal.capex.recurring_reserve_per_unit, "dollar"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1
    write_label(ws, f"B{r}", "Renovation Schedule (% units/yr)")
    for i, g in enumerate(deal.capex.units_renovated_pct):
        write_input(ws, (r, 4 + i), g, fmt="pct1")
    r += 2

    write_section(ws, r, "Debt"); r += 2
    for label, val, fmt in [
        ("Rate", deal.debt.rate, "pct2"),
        ("Term (yrs)", deal.debt.term_yrs, "general"),
        ("Amort (yrs)", deal.debt.amort_yrs, "general"),
        ("IO Period (yrs)", deal.debt.io_period_yrs, "general"),
        ("Max LTV", deal.debt.max_ltv, "pct1"),
        ("Min DSCR", deal.debt.min_dscr, "multiple"),
        ("Min Debt Yield", deal.debt.min_debt_yield, "pct1"),
        ("Origination Fee %", deal.debt.origination_fee_pct, "pct2"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1
    write_section(ws, r, "Equity / Waterfall"); r += 2
    for label, val, fmt in [
        ("Pref Rate", deal.equity.pref_rate, "pct1"),
        ("Promote %", deal.equity.promote_pct, "pct1"),
        ("GP Co-Invest %", deal.equity.gp_coinvest_pct, "pct1"),
        ("Acq Fee %", deal.equity.acq_fee_pct, "pct1"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1
    write_section(ws, r, "Exit"); r += 2
    for label, val, fmt in [
        ("Hold Period (yrs)", deal.exit.hold_yrs, "general"),
        ("Exit Cap Rate", deal.exit.exit_cap, "pct2"),
        ("Cost of Sale %", deal.exit.cost_of_sale_pct, "pct2"),
        ("Exit NOI Basis", deal.exit.exit_noi_basis, "general"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1


def _write_sources_uses(wb: Workbook, pf: ProForma):
    ws = wb.create_sheet("Sources & Uses")
    set_sheet_defaults(ws, "Sources & Uses")
    write_header(ws, 1, f"{pf.deal.deal_name} — Sources & Uses", span_cols=6)

    su = pf.sources_uses
    r = 3
    write_section(ws, r, "Uses"); r += 2
    rows_uses = [
        ("Purchase Price", su.purchase_price),
        ("Closing Costs", su.closing_costs),
        ("Initial CapEx (Day 1)", su.initial_capex),
        ("Day-One Reserves", su.day_one_reserves),
        ("Acquisition Fee", su.acq_fee),
        ("Origination Fee", su.origination_fee),
        ("Lender Reserves", su.lender_reserves),
    ]
    first_use_row = r
    for label, val in rows_uses:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt="dollar"); r += 1
    write_label(ws, f"B{r}", "Total Uses", bold=True)
    write_total(ws, f"D{r}", f"=SUM(D{first_use_row}:D{r-1})", fmt="dollar")
    total_uses_row = r
    r += 2

    write_section(ws, r, "Sources"); r += 2
    write_label(ws, f"B{r}", "Loan Amount"); write_input(ws, f"D{r}", su.loan_amount, fmt="dollar")
    loan_row = r; r += 1
    write_label(ws, f"B{r}", "Equity Check (LP + GP)")
    write_formula(ws, f"D{r}", f"=D{total_uses_row}-D{loan_row}", fmt="dollar")
    equity_row = r; r += 1
    write_label(ws, f"B{r}", "Total Sources", bold=True)
    write_total(ws, f"D{r}", f"=D{loan_row}+D{equity_row}", fmt="dollar")
    r += 2

    write_label(ws, f"B{r}", "All-In Basis / Unit")
    write_formula(ws, f"D{r}", f"=D{total_uses_row}/{pf.deal.property.unit_count}", fmt="dollar")
    return loan_row, equity_row


def _write_pro_forma(wb: Workbook, pf: ProForma):
    deal = pf.deal
    ws = wb.create_sheet("Pro Forma")
    set_sheet_defaults(ws, "Pro Forma")
    write_header(ws, 1, f"{deal.deal_name} — Pro Forma", span_cols=2 + len(pf.years) + 2)

    n = len(pf.years)
    year_labels = [f"Year {y.year}" for y in pf.years]
    _set_year_columns(ws, 4, n, year_labels)
    # Column widths for year cols
    for i in range(n):
        ws.column_dimensions[_col(4 + i)].width = 14

    r = 6
    write_section(ws, r, "Revenue"); r += 1
    # GPR
    write_label(ws, f"B{r}", "Gross Potential Rent"); write_units(ws, f"C{r}", "USD")
    for i, y in enumerate(pf.years):
        write_input(ws, (r, 4 + i), y.gpr, fmt="dollar")
    gpr_row = r; r += 1
    # Vacancy
    write_label(ws, f"B{r}", "(Vacancy)"); write_units(ws, f"C{r}", "USD")
    for i, y in enumerate(pf.years):
        write_input(ws, (r, 4 + i), y.vacancy, fmt="dollar")
    vac_row = r; r += 1
    # Concessions
    write_label(ws, f"B{r}", "(Concessions)"); write_units(ws, f"C{r}", "USD")
    for i, y in enumerate(pf.years):
        write_input(ws, (r, 4 + i), y.concessions, fmt="dollar")
    conc_row = r; r += 1
    # Bad debt
    write_label(ws, f"B{r}", "(Bad Debt)"); write_units(ws, f"C{r}", "USD")
    for i, y in enumerate(pf.years):
        write_input(ws, (r, 4 + i), y.bad_debt, fmt="dollar")
    bd_row = r; r += 1
    # Other income
    write_label(ws, f"B{r}", "Other Income"); write_units(ws, f"C{r}", "USD")
    for i, y in enumerate(pf.years):
        write_input(ws, (r, 4 + i), y.other_income, fmt="dollar")
    oi_row = r; r += 1
    # EGI subtotal (formula)
    write_label(ws, f"B{r}", "Effective Gross Income", bold=True)
    for i in range(n):
        c = _col(4 + i)
        write_subtotal(ws, (r, 4 + i), f"=SUM({c}{gpr_row}:{c}{oi_row})", fmt="dollar")
    egi_row = r; r += 2

    write_section(ws, r, "Operating Expenses"); r += 1
    opex_rows = []
    for label, attr in [
        ("Payroll", "payroll"), ("R&M", "rm"), ("Marketing", "marketing"),
        ("Utilities", "utilities"), ("Insurance", "insurance"),
        ("Other OpEx", "other_opex"), ("RE Tax", "re_tax"), ("Mgmt Fee", "mgmt_fee"),
    ]:
        write_label(ws, f"B{r}", label); write_units(ws, f"C{r}", "USD")
        for i, y in enumerate(pf.years):
            write_input(ws, (r, 4 + i), getattr(y, attr), fmt="dollar")
        opex_rows.append(r); r += 1
    write_label(ws, f"B{r}", "Total OpEx", bold=True)
    for i in range(n):
        c = _col(4 + i)
        write_subtotal(ws, (r, 4 + i), f"=SUM({c}{opex_rows[0]}:{c}{opex_rows[-1]})", fmt="dollar")
    opex_total_row = r; r += 2

    write_label(ws, f"B{r}", "Net Operating Income", bold=True)
    for i in range(n):
        c = _col(4 + i)
        write_total(ws, (r, 4 + i), f"={c}{egi_row}-{c}{opex_total_row}", fmt="dollar")
    noi_row = r; r += 2

    write_section(ws, r, "CapEx"); r += 1
    capex_rows = []
    for label, attr in [
        ("Value-Add CapEx", "value_add_capex"),
        ("Common Area CapEx", "common_area_capex"),
        ("Recurring Reserve", "recurring_reserve"),
    ]:
        write_label(ws, f"B{r}", label); write_units(ws, f"C{r}", "USD")
        for i, y in enumerate(pf.years):
            write_input(ws, (r, 4 + i), getattr(y, attr), fmt="dollar")
        capex_rows.append(r); r += 1
    write_label(ws, f"B{r}", "Total CapEx", bold=True)
    for i in range(n):
        c = _col(4 + i)
        write_subtotal(ws, (r, 4 + i), f"=SUM({c}{capex_rows[0]}:{c}{capex_rows[-1]})", fmt="dollar")
    capex_total_row = r; r += 2

    write_label(ws, f"B{r}", "Unlevered Net Cash Flow", bold=True)
    for i in range(n):
        c = _col(4 + i)
        write_total(ws, (r, 4 + i), f"={c}{noi_row}-{c}{capex_total_row}", fmt="dollar")
    ncf_unlev_row = r; r += 2

    write_section(ws, r, "Debt Service"); r += 1
    write_label(ws, f"B{r}", "Interest"); write_units(ws, f"C{r}", "USD")
    for i, y in enumerate(pf.years):
        write_input(ws, (r, 4 + i), -y.interest, fmt="dollar")
    int_row = r; r += 1
    write_label(ws, f"B{r}", "Principal"); write_units(ws, f"C{r}", "USD")
    for i, y in enumerate(pf.years):
        write_input(ws, (r, 4 + i), -y.principal, fmt="dollar")
    princ_row = r; r += 1
    write_label(ws, f"B{r}", "Total Debt Service", bold=True)
    for i in range(n):
        c = _col(4 + i)
        write_subtotal(ws, (r, 4 + i), f"={c}{int_row}+{c}{princ_row}", fmt="dollar")
    ds_row = r; r += 2

    write_label(ws, f"B{r}", "Levered Net Cash Flow", bold=True)
    for i in range(n):
        c = _col(4 + i)
        write_total(ws, (r, 4 + i), f"={c}{ncf_unlev_row}+{c}{ds_row}", fmt="dollar")
    ncf_lev_row = r; r += 1

    # Returns metrics inline
    r += 2
    write_subheader(ws, r, "Going-In Metrics"); r += 1
    write_label(ws, f"B{r}", "Going-In Cap")
    write_formula(ws, f"D{r}", f"=D{noi_row}/'Sources & Uses'!D5", fmt="pct2")
    r += 1
    write_label(ws, f"B{r}", "Year 3 Cap (stabilized)")
    if n >= 3:
        write_formula(ws, f"D{r}", f"=F{noi_row}/'Sources & Uses'!D5", fmt="pct2")

    ws.freeze_panes = "D6"
    return noi_row, ncf_unlev_row, ncf_lev_row


def _write_debt_sheet(wb: Workbook, pf: ProForma):
    ws = wb.create_sheet("Debt")
    set_sheet_defaults(ws, "Debt")
    write_header(ws, 1, f"{pf.deal.deal_name} — Debt Sizing & Schedule", span_cols=8)

    s = pf.sizing
    r = 3
    write_section(ws, r, "Sizing — Most-Binding Constraint"); r += 2
    for label, val in [
        ("Sizing NOI (Year 1)", pf.years[0].noi),
        ("Max via LTV", s.constraints.get("LTV", 0)),
        ("Max via DSCR", s.constraints.get("DSCR", 0)),
        ("Max via Debt Yield", s.constraints.get("Debt Yield", 0)),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt="dollar"); r += 1
    write_label(ws, f"B{r}", "Binding Constraint", bold=True); write_input(ws, f"D{r}", s.binding, fmt="general"); r += 1
    write_label(ws, f"B{r}", "Loan Amount", bold=True); write_total(ws, f"D{r}", s.loan_amount, fmt="dollar"); r += 1
    write_label(ws, f"B{r}", "Implied LTV"); write_input(ws, f"D{r}", s.implied_ltv, fmt="pct1"); r += 1
    write_label(ws, f"B{r}", "Implied DSCR"); write_input(ws, f"D{r}", s.implied_dscr, fmt="multiple"); r += 1
    write_label(ws, f"B{r}", "Implied Debt Yield"); write_input(ws, f"D{r}", s.implied_debt_yield, fmt="pct1"); r += 2

    write_section(ws, r, "Annual Amortization Schedule"); r += 1
    headers = ["Year", "Beg. Balance", "Interest", "Principal", "Debt Service", "End. Balance", "IO?"]
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=2 + i, value=h)
        c.style = "label_bold"
        c.alignment = c.alignment.copy(horizontal="right")
    r += 1
    for am in pf.amort_schedule:
        ws.cell(row=r, column=2, value=am.year).style = "label"
        for col_idx, val in enumerate([am.beginning_balance, am.interest, am.principal, am.debt_service, am.ending_balance], start=3):
            write_input(ws, (r, col_idx), val, fmt="dollar")
        ws.cell(row=r, column=8, value="IO" if am.is_io else "Amort").style = "label"
        r += 1


def _write_returns(wb: Workbook, pf: ProForma, wf: WaterfallResult):
    ws = wb.create_sheet("Returns")
    set_sheet_defaults(ws, "Returns")
    write_header(ws, 1, f"{pf.deal.deal_name} — Returns", span_cols=8)

    r = 3
    write_section(ws, r, "Equity Cash Flows (Total Equity, LP + GP)"); r += 2
    # Headers
    write_label(ws, f"B{r}", "Date", bold=True)
    write_label(ws, f"D{r}", "Cash Flow", bold=True)
    r += 1
    cf_first_row = r
    for ef in pf.equity_flows_total:
        write_input(ws, f"B{r}", ef.period, fmt="date")
        write_input(ws, f"D{r}", ef.amount, fmt="dollar")
        r += 1
    cf_last_row = r - 1
    r += 1

    write_section(ws, r, "Returns Summary"); r += 2

    # Total equity (project) — Excel XIRR formula using the date+amount range
    write_label(ws, f"B{r}", "Total Equity IRR (Project)", bold=True)
    write_total(ws, f"D{r}", f"=XIRR(D{cf_first_row}:D{cf_last_row},B{cf_first_row}:B{cf_last_row})", fmt="pct1")
    r += 1
    write_label(ws, f"B{r}", "Total Equity MOIC")
    write_input(ws, f"D{r}", wf.total_equity_moic, fmt="multiple")
    r += 2

    write_subheader(ws, r, "Net of Waterfall (LP / GP)"); r += 1
    write_label(ws, f"B{r}", "LP Net IRR"); write_input(ws, f"D{r}", wf.lp.irr, fmt="pct1"); r += 1
    write_label(ws, f"B{r}", "LP Net MOIC"); write_input(ws, f"D{r}", wf.lp.moic, fmt="multiple"); r += 1
    write_label(ws, f"B{r}", "LP Contributed"); write_input(ws, f"D{r}", wf.lp.contributed, fmt="dollar"); r += 1
    write_label(ws, f"B{r}", "LP Distributed"); write_input(ws, f"D{r}", wf.lp.distributed, fmt="dollar"); r += 2

    write_label(ws, f"B{r}", "GP Net IRR (co-invest + promote)"); write_input(ws, f"D{r}", wf.gp.irr, fmt="pct1"); r += 1
    write_label(ws, f"B{r}", "GP Net MOIC"); write_input(ws, f"D{r}", wf.gp.moic, fmt="multiple"); r += 1
    write_label(ws, f"B{r}", "GP Contributed"); write_input(ws, f"D{r}", wf.gp.contributed, fmt="dollar"); r += 1
    write_label(ws, f"B{r}", "GP Distributed"); write_input(ws, f"D{r}", wf.gp.distributed, fmt="dollar"); r += 2

    write_subheader(ws, r, "Exit Reversion"); r += 1
    e = pf.exit_summary
    for label, val, fmt in [
        ("Exit NOI", e.exit_noi, "dollar"),
        ("Exit Cap", e.exit_cap, "pct2"),
        ("Gross Sale Price", e.gross_sale, "dollar"),
        ("(Cost of Sale)", -e.cost_of_sale, "dollar"),
        ("(Loan Payoff)", -e.loan_payoff, "dollar"),
        ("Net Proceeds to Equity", e.net_proceeds, "dollar"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    write_note(ws, r + 1, f"Waterfall: {wf.pref_rate*100:.1f}% pref, {wf.promote_pct*100:.0f}% promote (American). LP/GP MOIC and IRR shown after promote.")


def _write_sensitivities(wb: Workbook, tables: list[SensitivityTable]):
    ws = wb.create_sheet("Sensitivities")
    set_sheet_defaults(ws, "Sensitivities")
    write_header(ws, 1, "Sensitivity Tables", span_cols=8)

    r = 3
    for table in tables:
        write_section(ws, r, table.title); r += 1
        write_subheader(ws, r, f"{table.y_label}  vs.  {table.x_label}"); r += 2

        # Column headers
        write_label(ws, f"C{r}", table.x_label, bold=True)
        for i, xv in enumerate(table.x_values):
            c = ws.cell(row=r, column=4 + i, value=xv)
            if "Cap" in table.x_label or "Growth" in table.x_label or "LTV" in table.x_label or "Δ" in table.x_label:
                fmt = "pct2"
            else:
                fmt = "general"
            c.style = f"input_{fmt}"
        r += 1

        # Rows
        is_pct_metric = "IRR" in table.metric or "MOIC" not in table.metric
        cell_fmt = "pct1" if "IRR" in table.metric else "multiple"
        for yi, yv in enumerate(table.y_values):
            label_yv = ws.cell(row=r, column=3, value=yv)
            label_yv.style = "input_pct1" if "Growth" not in table.y_label else "input_multiple"
            for xi, val in enumerate(table.cells[yi]):
                write_input(ws, (r, 4 + xi), val, fmt=cell_fmt)
            r += 1
        r += 2


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_workbook(
    pf: ProForma,
    wf: WaterfallResult,
    sensitivities: list[SensitivityTable],
    out_path: str | Path,
) -> Path:
    wb = Workbook()
    apply_institutional_styles(wb)
    # Drop the default sheet
    default = wb.active
    wb.remove(default)

    _write_assumptions(wb, pf)
    _write_sources_uses(wb, pf)
    _write_pro_forma(wb, pf)
    _write_debt_sheet(wb, pf)
    _write_returns(wb, pf, wf)
    _write_sensitivities(wb, sensitivities)

    # v5 institutional Executive Summary as first tab
    from .excel_summary import build_payload, write_executive_summary
    value_add_total = (
        pf.deal.capex.value_add_per_unit * pf.deal.property.unit_count
        if pf.deal.capex.value_add_per_unit > 0 else 0.0
    )
    payload = build_payload(
        pf=pf, wf=wf,
        asset_class="Multifamily",
        denom_label="Units",
        denom_value=pf.deal.property.unit_count,
        per_denom_label="$/Unit",
        per_denom_fmt="per_unit",
        value_add_capex_total=value_add_total,
    )
    write_executive_summary(wb, payload)

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
