"""
excel_writer.py — Institutional XLSX output for hospitality underwriting.

Sheets:
  - Assumptions       property + operating + opex + capex + debt + equity + exit
  - Operating Stats   ADR / Occ / RevPAR / sold room-nights / displacement by year
  - Departmental P&L  Rooms / F&B / Other revenue, expense, profit
  - Pro Forma         USALI rollup: revenue -> dept profit -> GOP -> NOI -> NCF
  - Debt              sizing + amort schedule
  - Returns           XIRR/MOIC + waterfall + S&U + exit
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

import scripts  # noqa: F401
from excel_style import (
    apply_institutional_styles,
    set_sheet_defaults,
    write_formula,
    write_header,
    write_input,
    write_label,
    write_section,
    write_subtotal,
)

from ..waterfall_acq import WaterfallResult
from .pro_forma import HotelProForma


# ---------------------------------------------------------------------------
# Assumptions
# ---------------------------------------------------------------------------

def _write_assumptions(wb: Workbook, pf: HotelProForma):
    deal = pf.deal
    ws = wb.create_sheet("Assumptions")
    set_sheet_defaults(ws, f"{deal.deal_name} - Assumptions")
    write_header(ws, 1, f"{deal.deal_name} - Assumptions", span_cols=8)

    r = 3
    write_section(ws, r, "Property"); r += 2
    for label, val, fmt in [
        ("Asset Class", deal.property.asset_class, "general"),
        ("Submarket", deal.property.submarket, "general"),
        ("Brand", deal.property.brand, "general"),
        ("Flag Type", deal.property.flag_type, "general"),
        ("Service Level", deal.property.service_level, "general"),
        ("Year Built", deal.property.year_built, "general"),
        ("Keys", deal.property.keys, "general"),
        ("Available Room-Nights / Yr", deal.property.available_room_nights, "general"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Acquisition"); r += 2
    for label, val, fmt in [
        ("Purchase Price", deal.acquisition.purchase_price, "dollar"),
        ("Price / Key", deal.acquisition.purchase_price / deal.property.keys, "per_unit"),
        ("Closing Costs %", deal.acquisition.closing_costs_pct, "pct2"),
        ("Initial CapEx (Day 1)", deal.acquisition.initial_capex, "dollar"),
        ("Day-One Reserves", deal.acquisition.day_one_reserves, "dollar"),
        ("Close Date", deal.acquisition.close_date, "date"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Operating Assumptions"); r += 2
    op = deal.operating
    for label, val, fmt in [
        ("Yr 1 ADR", op.adr_yr1, "dollar"),
        ("ADR Growth", op.adr_growth, "pct1"),
        ("Rooms Expense % of Rooms Rev", op.rooms_expense_pct, "pct1"),
        ("F&B Revenue % of Rooms Rev", op.fb_revenue_pct_of_rooms, "pct1"),
        ("F&B Margin", op.fb_margin, "pct1"),
        ("Other Revenue % of Rooms Rev", op.other_revenue_pct_of_rooms, "pct1"),
        ("Other Margin", op.other_margin, "pct1"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    write_label(ws, f"B{r}", "Occupancy by Year"); r += 1
    for i, occ in enumerate(op.occupancy, start=1):
        write_label(ws, f"B{r}", f"  Yr {i}"); write_input(ws, f"D{r}", occ, fmt="pct1"); r += 1

    r += 1; write_section(ws, r, "OpEx (Undistributed / Fixed)"); r += 2
    for label, val, fmt in [
        ("G&A (PAR)", deal.opex.ga_par, "per_unit"),
        ("S&M (PAR)", deal.opex.sm_par, "per_unit"),
        ("R&M (PAR)", deal.opex.rm_par, "per_unit"),
        ("Utilities (PAR)", deal.opex.utilities_par, "per_unit"),
        ("Insurance (PAR)", deal.opex.insurance_par, "per_unit"),
        ("RE Tax (annual)", deal.opex.re_tax, "dollar"),
        ("Franchise Fee % of Rooms Rev", deal.opex.franchise_fee_pct, "pct1"),
        ("Mgmt Fee % of Total Rev", deal.opex.mgmt_fee_pct, "pct1"),
        ("FF&E Reserve % of Total Rev", deal.opex.ffe_reserve_pct, "pct1"),
        ("Undistributed Growth", deal.opex.undistributed_growth, "pct1"),
        ("RE Tax Growth", deal.opex.re_tax_growth, "pct1"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "PIP / CapEx"); r += 2
    for label, val, fmt in [
        ("PIP Total", deal.capex.pip_total, "dollar"),
        ("PIP / Key", deal.capex.pip_total / deal.property.keys if deal.property.keys else 0, "per_unit"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1
    if deal.capex.pip_schedule_pct:
        write_label(ws, f"B{r}", "PIP Schedule by Year"); r += 1
        for i, pct in enumerate(deal.capex.pip_schedule_pct, start=1):
            disp = deal.capex.pip_displacement_keys[i - 1] if i - 1 < len(deal.capex.pip_displacement_keys) else 0
            write_label(ws, f"B{r}", f"  Yr {i}")
            write_input(ws, f"D{r}", pct, fmt="pct1")
            write_label(ws, f"F{r}", "Displaced Keys")
            write_input(ws, f"H{r}", disp, fmt="general")
            r += 1

    r += 1; write_section(ws, r, "Debt"); r += 2
    for label, val, fmt in [
        ("Rate", deal.debt.rate, "pct2"),
        ("Term (yrs)", deal.debt.term_yrs, "general"),
        ("Amort (yrs)", deal.debt.amort_yrs, "general"),
        ("IO Period (yrs)", deal.debt.io_period_yrs, "general"),
        ("Max LTV", deal.debt.max_ltv, "pct1"),
        ("Min DSCR", deal.debt.min_dscr, "multiple"),
        ("Min Debt Yield", deal.debt.min_debt_yield, "pct1"),
        ("Origination Fee %", deal.debt.origination_fee_pct, "pct2"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Equity / Waterfall"); r += 2
    for label, val, fmt in [
        ("Pref Rate", deal.equity.pref_rate, "pct1"),
        ("Promote %", deal.equity.promote_pct, "pct1"),
        ("GP Co-Invest %", deal.equity.gp_coinvest_pct, "pct1"),
        ("Acq Fee %", deal.equity.acq_fee_pct, "pct1"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Exit"); r += 2
    for label, val, fmt in [
        ("Hold Period (yrs)", deal.exit.hold_yrs, "general"),
        ("Exit Cap Rate", deal.exit.exit_cap, "pct2"),
        ("Cost of Sale %", deal.exit.cost_of_sale_pct, "pct2"),
        ("Exit NOI Basis", deal.exit.exit_noi_basis, "general"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1


# ---------------------------------------------------------------------------
# Year-column helper
# ---------------------------------------------------------------------------

def _year_headers(ws, n_years: int, row: int = 3, start_col: int = 3):
    ws.cell(row=row, column=2, value="").style = "label_bold"
    for y in range(1, n_years + 1):
        c = ws.cell(row=row, column=start_col + y - 1, value=f"Yr {y}")
        c.style = "label_bold"
        c.alignment = c.alignment.copy(horizontal="right")


# ---------------------------------------------------------------------------
# Operating Stats
# ---------------------------------------------------------------------------

def _write_operating_stats(wb: Workbook, pf: HotelProForma):
    ws = wb.create_sheet("Operating Stats")
    set_sheet_defaults(ws, "Operating Stats")
    write_header(ws, 1, f"{pf.deal.deal_name} - Operating Stats", span_cols=10)

    n = len(pf.years)
    _year_headers(ws, n)

    r = 4
    def row(label: str, getter, fmt: str = "dollar", bold: bool = False):
        nonlocal r
        write_label(ws, f"B{r}", label, bold=bold)
        for i, yl in enumerate(pf.years):
            write_input(ws, (r, 3 + i), getter(yl), fmt=fmt)
        r += 1

    row("ADR", lambda y: y.adr)
    row("Occupancy", lambda y: y.occupancy, fmt="pct1")
    row("Avg Keys Available", lambda y: y.avg_keys_available, fmt="general")
    row("Sold Room-Nights", lambda y: y.sold_room_nights, fmt="general")
    row("RevPAR", lambda y: y.revpar)


# ---------------------------------------------------------------------------
# Departmental P&L
# ---------------------------------------------------------------------------

def _write_departmental(wb: Workbook, pf: HotelProForma):
    ws = wb.create_sheet("Departmental P&L")
    set_sheet_defaults(ws, "Departmental P&L")
    write_header(ws, 1, f"{pf.deal.deal_name} - Departmental P&L", span_cols=10)

    n = len(pf.years)
    _year_headers(ws, n)

    r = 4
    def row(label: str, getter, fmt: str = "dollar", bold: bool = False):
        nonlocal r
        write_label(ws, f"B{r}", label, bold=bold)
        for i, yl in enumerate(pf.years):
            write_input(ws, (r, 3 + i), getter(yl), fmt=fmt)
        r += 1

    write_section(ws, r, "Revenue"); r += 2
    row("Rooms", lambda y: y.rooms_revenue)
    row("F&B", lambda y: y.fb_revenue)
    row("Other", lambda y: y.other_revenue)
    write_label(ws, f"B{r}", "Total Revenue", bold=True)
    for i in range(n):
        col_letter = ws.cell(row=r, column=3 + i).column_letter
        write_subtotal(ws, (r, 3 + i), f"=SUM({col_letter}{r-3}:{col_letter}{r-1})", fmt="dollar")
    r += 2

    write_section(ws, r, "Departmental Expense"); r += 2
    row("Rooms Expense", lambda y: -y.rooms_expense)
    row("F&B Expense", lambda y: -y.fb_expense)
    row("Other Expense", lambda y: -y.other_expense)
    write_label(ws, f"B{r}", "Total Dept Expense", bold=True)
    for i in range(n):
        col_letter = ws.cell(row=r, column=3 + i).column_letter
        write_subtotal(ws, (r, 3 + i), f"=SUM({col_letter}{r-3}:{col_letter}{r-1})", fmt="dollar")
    r += 2

    write_label(ws, f"B{r}", "Departmental Profit", bold=True)
    for i in range(n):
        write_input(ws, (r, 3 + i), pf.years[i].total_dept_profit, fmt="dollar")
    r += 2

    write_label(ws, f"B{r}", "Rooms GOP Margin")
    for i in range(n):
        yl = pf.years[i]
        margin = (yl.rooms_revenue - yl.rooms_expense) / yl.rooms_revenue if yl.rooms_revenue else 0.0
        write_input(ws, (r, 3 + i), margin, fmt="pct1")


# ---------------------------------------------------------------------------
# Pro Forma (USALI rollup)
# ---------------------------------------------------------------------------

def _write_pro_forma(wb: Workbook, pf: HotelProForma):
    ws = wb.create_sheet("Pro Forma")
    set_sheet_defaults(ws, "Pro Forma")
    write_header(ws, 1, f"{pf.deal.deal_name} - Pro Forma (USALI)", span_cols=10)

    n = len(pf.years)
    _year_headers(ws, n)

    r = 4
    def row(label: str, getter, fmt: str = "dollar"):
        nonlocal r
        write_label(ws, f"B{r}", label)
        for i, yl in enumerate(pf.years):
            write_input(ws, (r, 3 + i), getter(yl), fmt=fmt)
        r += 1

    write_label(ws, f"B{r}", "Total Revenue", bold=True)
    for i in range(n):
        write_input(ws, (r, 3 + i), pf.years[i].total_revenue, fmt="dollar")
    r += 1
    row("Less: Total Dept Expense", lambda y: -y.total_dept_expense)
    write_label(ws, f"B{r}", "Departmental Profit", bold=True)
    for i in range(n):
        write_input(ws, (r, 3 + i), pf.years[i].total_dept_profit, fmt="dollar")
    r += 2

    write_section(ws, r, "Undistributed"); r += 2
    row("G&A", lambda y: -y.ga)
    row("S&M", lambda y: -y.sm)
    row("R&M", lambda y: -y.rm)
    row("Utilities", lambda y: -y.utilities)
    row("Franchise Fee", lambda y: -y.franchise_fee)
    write_label(ws, f"B{r}", "Total Undistributed", bold=True)
    for i in range(n):
        col_letter = ws.cell(row=r, column=3 + i).column_letter
        write_subtotal(ws, (r, 3 + i), f"=SUM({col_letter}{r-5}:{col_letter}{r-1})", fmt="dollar")
    r += 2

    write_label(ws, f"B{r}", "GOP", bold=True)
    for i in range(n):
        write_input(ws, (r, 3 + i), pf.years[i].gop, fmt="dollar")
    r += 1
    row("Less: Mgmt Fee", lambda y: -y.mgmt_fee)
    write_label(ws, f"B{r}", "Income Before Fixed Charges", bold=True)
    for i in range(n):
        write_input(ws, (r, 3 + i), pf.years[i].ibfc, fmt="dollar")
    r += 2

    write_section(ws, r, "Fixed Charges"); r += 2
    row("RE Tax", lambda y: -y.re_tax)
    row("Insurance", lambda y: -y.insurance)
    write_label(ws, f"B{r}", "Total Fixed Charges", bold=True)
    for i in range(n):
        col_letter = ws.cell(row=r, column=3 + i).column_letter
        write_subtotal(ws, (r, 3 + i), f"=SUM({col_letter}{r-2}:{col_letter}{r-1})", fmt="dollar")
    r += 2

    write_label(ws, f"B{r}", "NOI (Pre-Reserve)", bold=True)
    for i in range(n):
        write_input(ws, (r, 3 + i), pf.years[i].noi_pre_reserve, fmt="dollar")
    r += 1
    row("Less: FF&E Reserve", lambda y: -y.ffe_reserve)
    write_label(ws, f"B{r}", "NOI (Cap-Rate Basis)", bold=True)
    for i in range(n):
        write_input(ws, (r, 3 + i), pf.years[i].noi, fmt="dollar")
    r += 2

    write_section(ws, r, "Cash Flow"); r += 2
    row("Less: PIP CapEx", lambda y: -y.pip_capex)
    write_label(ws, f"B{r}", "NCF Unlevered", bold=True)
    for i in range(n):
        write_input(ws, (r, 3 + i), pf.years[i].ncf_unlevered, fmt="dollar")
    r += 1
    row("Less: Debt Service", lambda y: -y.debt_service)
    write_label(ws, f"B{r}", "NCF Levered", bold=True)
    for i in range(n):
        write_input(ws, (r, 3 + i), pf.years[i].ncf_levered, fmt="dollar")


# ---------------------------------------------------------------------------
# Debt
# ---------------------------------------------------------------------------

def _write_debt(wb: Workbook, pf: HotelProForma):
    ws = wb.create_sheet("Debt")
    set_sheet_defaults(ws, "Debt Sizing")
    write_header(ws, 1, f"{pf.deal.deal_name} - Debt Sizing & Amort", span_cols=8)

    r = 3
    write_section(ws, r, "Sizing (Yr 1 NOI, post-reserve)"); r += 2
    s = pf.sizing
    for label, val, fmt in [
        ("Loan Amount", s.loan_amount, "dollar"),
        ("Binding Constraint", s.binding, "general"),
        ("Max LTV Capacity", s.constraints.get("LTV", 0), "dollar"),
        ("DSCR Capacity", s.constraints.get("DSCR", 0), "dollar"),
        ("Debt Yield Capacity", s.constraints.get("Debt Yield", 0), "dollar"),
        ("Implied LTV", s.implied_ltv, "pct1"),
        ("Implied DSCR", s.implied_dscr, "multiple"),
        ("Implied Debt Yield", s.implied_debt_yield, "pct2"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Amortization Schedule"); r += 2
    headers = ["Yr", "Beg. Balance", "Interest", "Principal", "Debt Service", "End Balance", "IO?"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=r, column=i, value=h).style = "label_bold"
    r += 1
    for am in pf.amort_schedule:
        ws.cell(row=r, column=2, value=am.year)
        write_input(ws, (r, 3), am.beginning_balance, fmt="dollar")
        write_input(ws, (r, 4), am.interest, fmt="dollar")
        write_input(ws, (r, 5), am.principal, fmt="dollar")
        write_input(ws, (r, 6), am.debt_service, fmt="dollar")
        write_input(ws, (r, 7), am.ending_balance, fmt="dollar")
        ws.cell(row=r, column=8, value="Yes" if am.is_io else "No")
        r += 1


# ---------------------------------------------------------------------------
# Returns
# ---------------------------------------------------------------------------

def _write_returns(wb: Workbook, pf: HotelProForma, wf: WaterfallResult):
    ws = wb.create_sheet("Returns")
    set_sheet_defaults(ws, "Returns")
    write_header(ws, 1, f"{pf.deal.deal_name} - Returns", span_cols=8)

    r = 3
    write_section(ws, r, "Sources & Uses"); r += 2
    su = pf.sources_uses
    for label, val, fmt in [
        ("Purchase Price", su.purchase_price, "dollar"),
        ("Closing Costs", su.closing_costs, "dollar"),
        ("Initial CapEx", su.initial_capex, "dollar"),
        ("Day-One Reserves", su.day_one_reserves, "dollar"),
        ("Acq Fee", su.acq_fee, "dollar"),
        ("Origination Fee", su.origination_fee, "dollar"),
        ("Lender Reserves", su.lender_reserves, "dollar"),
        ("Total Uses", su.total_uses, "dollar"),
        ("Loan Amount", su.loan_amount, "dollar"),
        ("Equity Check", su.equity_check, "dollar"),
        ("  per Key", pf.all_in_basis_per_key, "per_unit"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Exit Summary"); r += 2
    ex = pf.exit_summary
    for label, val, fmt in [
        ("Exit Year", ex.exit_year, "general"),
        ("Exit NOI (post-reserve)", ex.exit_noi, "dollar"),
        ("Exit Cap", ex.exit_cap, "pct2"),
        ("Gross Sale Price", ex.gross_sale, "dollar"),
        ("Cost of Sale", ex.cost_of_sale, "dollar"),
        ("Loan Payoff", ex.loan_payoff, "dollar"),
        ("Net Proceeds to Equity", ex.net_proceeds, "dollar"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Returns"); r += 2
    for label, val, fmt in [
        ("Going-In Cap (post-reserve)", pf.going_in_cap, "pct2"),
        ("Yr 3 Stabilized Cap", pf.stabilized_cap, "pct2"),
        ("Total Equity IRR (project)", wf.total_equity_irr, "pct2"),
        ("Total Equity MOIC", wf.total_equity_moic, "multiple"),
        ("LP Net IRR", wf.lp.irr, "pct2"),
        ("LP Net MOIC", wf.lp.moic, "multiple"),
        ("GP Net IRR", wf.gp.irr, "pct2"),
        ("GP Net MOIC", wf.gp.moic, "multiple"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Equity Cash Flows"); r += 2
    write_label(ws, f"B{r}", "Date", bold=True)
    write_label(ws, f"D{r}", "Amount", bold=True); r += 1
    for ef in pf.equity_flows_total:
        write_input(ws, (r, 2), ef.period, fmt="date")
        write_input(ws, (r, 4), ef.amount, fmt="dollar")
        r += 1


# ---------------------------------------------------------------------------
# Top-level
# ---------------------------------------------------------------------------

def write_hotel_workbook(pf: HotelProForma, wf: WaterfallResult, out_path: Path) -> Path:
    """Write the full hospitality workbook and return the path."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    apply_institutional_styles(wb)

    _write_assumptions(wb, pf)
    _write_operating_stats(wb, pf)
    _write_departmental(wb, pf)
    _write_pro_forma(wb, pf)
    _write_debt(wb, pf)
    _write_returns(wb, pf, wf)

    from ..excel_summary import build_payload, write_executive_summary
    payload = build_payload(
        pf=pf, wf=wf,
        asset_class="Hospitality",
        denom_label="Keys",
        denom_value=pf.deal.property.keys,
        per_denom_label="$/Key",
        per_denom_fmt="per_unit",
        value_add_capex_total=pf.deal.capex.pip_total,
    )
    write_executive_summary(wb, payload)

    wb.save(out_path)
    return out_path
