"""
excel_writer.py — Institutional XLSX output for commercial underwriting.

Sheets:
  - Assumptions     property + market + opex + debt + equity + exit inputs
  - Rent Roll       lease detail (tenant, SF, $/SF, type, dates, escalations)
  - Per-Lease CFs   year-by-year CF for each tenant (rent / free / recoveries / TI / LC)
  - Pro Forma       Yr 1..hold[+1] property-level P&L with SUM formulas
  - Debt            sizing + amort schedule
  - Returns         XIRR/MOIC + waterfall outputs
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

import scripts  # noqa: F401
from excel_style import (
    apply_institutional_styles,
    set_sheet_defaults,
    write_formula,
    write_header,
    write_input,
    write_label,
    write_section,
    write_subtotal,
    write_total,
)

from ..waterfall_acq import WaterfallResult
from .pro_forma import CommercialProForma


# ---------------------------------------------------------------------------
# Assumptions
# ---------------------------------------------------------------------------

def _write_assumptions(wb: Workbook, pf: CommercialProForma):
    deal = pf.deal
    ws = wb.create_sheet("Assumptions")
    set_sheet_defaults(ws, f"{deal.deal_name} — Assumptions")
    write_header(ws, 1, f"{deal.deal_name} — Assumptions", span_cols=8)

    r = 3
    write_section(ws, r, "Property"); r += 2
    for label, val, fmt in [
        ("Asset Class", deal.property.asset_class, "general"),
        ("Submarket", deal.property.submarket, "general"),
        ("Year Built", deal.property.year_built, "general"),
        ("Total RBA (SF)", deal.property.total_rba, "general"),
        ("Leased SF", deal.property.leased_sf, "general"),
        ("In-Place Occupancy", deal.property.in_place_occupancy, "pct1"),
        ("In-Place Gross Rent ($/yr)", deal.property.in_place_gross_rent, "dollar"),
        ("General Vacancy Reserve", deal.property.general_vacancy_pct, "pct1"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Acquisition"); r += 2
    for label, val, fmt in [
        ("Purchase Price", deal.acquisition.purchase_price, "dollar"),
        ("Price / SF", deal.acquisition.purchase_price / deal.property.total_rba, "per_sf"),
        ("Closing Costs %", deal.acquisition.closing_costs_pct, "pct2"),
        ("Initial CapEx (Day 1)", deal.acquisition.initial_capex, "dollar"),
        ("Day-One Reserves", deal.acquisition.day_one_reserves, "dollar"),
        ("Close Date", deal.acquisition.close_date, "date"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Market (Re-Leasing)"); r += 2
    for label, val, fmt in [
        ("Market Rent ($/SF/yr)", deal.market.market_rent_psf, "per_sf"),
        ("Market Rent Growth", deal.market.market_rent_growth, "pct1"),
        ("Renewal Probability", deal.market.renewal_prob, "pct1"),
        ("Downtime (mo, new tenant)", deal.market.downtime_mo, "general"),
        ("New TI / SF", deal.market.new_ti_psf, "per_sf"),
        ("New LC %", deal.market.new_lc_pct, "pct1"),
        ("New Free Rent (mo)", deal.market.new_free_rent_mo, "general"),
        ("New Term (yrs)", deal.market.new_lease_term_yrs, "general"),
        ("New Escalation", deal.market.new_escalation_pct, "pct1"),
        ("Renewal TI / SF", deal.market.renewal_ti_psf, "per_sf"),
        ("Renewal LC %", deal.market.renewal_lc_pct, "pct1"),
        ("Renewal Free Rent (mo)", deal.market.renewal_free_rent_mo, "general"),
        ("Renewal Term (yrs)", deal.market.renewal_lease_term_yrs, "general"),
        ("Renewal Escalation", deal.market.renewal_escalation_pct, "pct1"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "OpEx"); r += 2
    for label, val, fmt in [
        ("CAM / SF", deal.opex.cam_psf, "per_sf"),
        ("RE Tax (annual)", deal.opex.re_tax, "dollar"),
        ("Insurance / SF", deal.opex.insurance_psf, "per_sf"),
        ("Utilities / SF (recoverable)", deal.opex.utilities_psf, "per_sf"),
        ("Non-Recoverable / SF", deal.opex.non_recoverable_psf, "per_sf"),
        ("Mgmt Fee % of EGI", deal.opex.mgmt_fee_pct, "pct1"),
        ("CAM Growth", deal.opex.cam_growth, "pct1"),
        ("RE Tax Growth", deal.opex.re_tax_growth, "pct1"),
        ("Insurance Growth", deal.opex.insurance_growth, "pct1"),
        ("Utilities Growth", deal.opex.utilities_growth, "pct1"),
        ("Non-Recoverable Growth", deal.opex.non_recoverable_growth, "pct1"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "CapEx"); r += 2
    for label, val, fmt in [
        ("Initial Building CapEx (Yr 1)", deal.capex.initial_building_capex, "dollar"),
        ("Recurring Reserve / SF", deal.capex.recurring_reserve_psf, "per_sf"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Debt"); r += 2
    for label, val, fmt in [
        ("Rate", deal.debt.rate, "pct2"),
        ("Term (yrs)", deal.debt.term_yrs, "general"),
        ("Amort (yrs)", deal.debt.amort_yrs, "general"),
        ("IO Period (yrs)", deal.debt.io_period_yrs, "general"),
        ("Max LTV", deal.debt.max_ltv, "pct1"),
        ("Min DSCR", deal.debt.min_dscr, "multiple"),
        ("Min Debt Yield", deal.debt.min_debt_yield, "pct1"),
        ("Origination Fee %", deal.debt.origination_fee_pct, "pct2"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Equity / Waterfall"); r += 2
    for label, val, fmt in [
        ("Pref Rate", deal.equity.pref_rate, "pct1"),
        ("Promote %", deal.equity.promote_pct, "pct1"),
        ("GP Co-Invest %", deal.equity.gp_coinvest_pct, "pct1"),
        ("Acq Fee %", deal.equity.acq_fee_pct, "pct1"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Exit"); r += 2
    for label, val, fmt in [
        ("Hold Period (yrs)", deal.exit.hold_yrs, "general"),
        ("Exit Cap Rate", deal.exit.exit_cap, "pct2"),
        ("Cost of Sale %", deal.exit.cost_of_sale_pct, "pct2"),
        ("Exit NOI Basis", deal.exit.exit_noi_basis, "general"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1


# ---------------------------------------------------------------------------
# Rent Roll
# ---------------------------------------------------------------------------

def _write_rent_roll(wb: Workbook, pf: CommercialProForma):
    deal = pf.deal
    ws = wb.create_sheet("Rent Roll")
    set_sheet_defaults(ws, "Rent Roll")
    write_header(ws, 1, f"{deal.deal_name} — Rent Roll", span_cols=10)

    headers = ["Tenant", "Suite", "SF", "$/SF/yr", "Annual Rent", "Type",
               "Lease Start", "Lease End", "Escalation %", "Free Rent (mo)"]
    for i, h in enumerate(headers, start=2):
        c = ws.cell(row=3, column=i, value=h); c.style = "label_bold"

    r = 4
    for lease in deal.property.rent_roll:
        ws.cell(row=r, column=2, value=lease.tenant)
        ws.cell(row=r, column=3, value=lease.suite or "")
        write_input(ws, (r, 4), lease.sf, fmt="general")
        write_input(ws, (r, 5), lease.base_rent_psf, fmt="per_sf")
        write_formula(ws, (r, 6), f"=D{r}*E{r}", fmt="dollar")
        ws.cell(row=r, column=7, value=lease.lease_type)
        write_input(ws, (r, 8), lease.lease_start, fmt="date")
        write_input(ws, (r, 9), lease.lease_end, fmt="date")
        write_input(ws, (r, 10), lease.escalation_pct, fmt="pct1")
        write_input(ws, (r, 11), lease.free_rent_remaining_mo, fmt="general")
        r += 1

    # Totals row
    write_label(ws, f"B{r}", "TOTAL", bold=True)
    write_formula(ws, f"D{r}", f"=SUM(D4:D{r-1})", fmt="sf")
    write_subtotal(ws, f"F{r}", f"=SUM(F4:F{r-1})", fmt="dollar")


# ---------------------------------------------------------------------------
# Per-Lease CFs
# ---------------------------------------------------------------------------

def _write_per_lease(wb: Workbook, pf: CommercialProForma):
    ws = wb.create_sheet("Per-Lease CFs")
    set_sheet_defaults(ws, "Per-Lease Cash Flows")
    write_header(ws, 1, f"{pf.deal.deal_name} — Per-Lease Cash Flows (probability-weighted)", span_cols=10)

    n_years = len(pf.years)
    # Year headers (row 3)
    ws.cell(row=3, column=2, value="Tenant").style = "label_bold"
    ws.cell(row=3, column=3, value="Line").style = "label_bold"
    for y in range(1, n_years + 1):
        c = ws.cell(row=3, column=3 + y, value=f"Yr {y}")
        c.style = "label_bold"; c.alignment = c.alignment.copy(horizontal="right")

    r = 4
    for tenant_key, years in pf.per_lease_years.items():
        write_label(ws, f"B{r}", tenant_key, bold=True); r += 1
        for line_name, getter in [
            ("Base Rent", lambda ly: ly.base_rent),
            ("Free Rent", lambda ly: ly.free_rent),
            ("Recoveries", lambda ly: ly.recoveries),
            ("Pct Rent", lambda ly: ly.pct_rent),
            ("TI", lambda ly: -ly.ti),
            ("LC", lambda ly: -ly.lc),
            ("Occ. Mo.", lambda ly: ly.occupied_months),
        ]:
            ws.cell(row=r, column=3, value=line_name)
            for i, ly in enumerate(years):
                fmt = "general" if line_name == "Occ. Mo." else "dollar"
                write_input(ws, (r, 4 + i), getter(ly), fmt=fmt)
            r += 1
        r += 1


# ---------------------------------------------------------------------------
# Pro Forma
# ---------------------------------------------------------------------------

def _write_pro_forma(wb: Workbook, pf: CommercialProForma):
    ws = wb.create_sheet("Pro Forma")
    set_sheet_defaults(ws, "Pro Forma")
    write_header(ws, 1, f"{pf.deal.deal_name} — Pro Forma", span_cols=10)

    n_years = len(pf.years)
    ws.cell(row=3, column=2, value="").style = "label_bold"
    for y in range(1, n_years + 1):
        c = ws.cell(row=3, column=2 + y, value=f"Yr {y}")
        c.style = "label_bold"; c.alignment = c.alignment.copy(horizontal="right")

    def row(label: str, getter, fmt: str = "dollar", bold: bool = False):
        nonlocal r
        write_label(ws, f"B{r}", label, bold=bold)
        for i, yl in enumerate(pf.years):
            write_input(ws, (r, 3 + i), getter(yl), fmt=fmt)
        r += 1

    r = 4
    write_section(ws, r, "Revenue"); r += 2
    row("Gross Rent", lambda y: y.gross_rent)
    row("Free Rent", lambda y: y.free_rent)
    row("Recoveries", lambda y: y.recoveries)
    row("Percentage Rent", lambda y: y.pct_rent)
    row("General Vacancy", lambda y: y.general_vacancy)
    # EGI = SUM of above 5 rows (rows r-5..r-1)
    write_label(ws, f"B{r}", "EGI", bold=True)
    for i in range(n_years):
        col_letter = ws.cell(row=r, column=3 + i).column_letter
        write_subtotal(ws, (r, 3 + i), f"=SUM({col_letter}{r-5}:{col_letter}{r-1})", fmt="dollar")
    r += 1

    r += 1; write_section(ws, r, "OpEx"); r += 2
    row("Recoverable Pool", lambda y: -y.recoverable_opex)
    row("Non-Recoverable", lambda y: -y.non_recoverable_opex)
    row("Mgmt Fee", lambda y: -y.mgmt_fee)
    write_label(ws, f"B{r}", "Total OpEx", bold=True)
    for i in range(n_years):
        col_letter = ws.cell(row=r, column=3 + i).column_letter
        write_subtotal(ws, (r, 3 + i), f"=SUM({col_letter}{r-3}:{col_letter}{r-1})", fmt="dollar")
    r += 2

    write_label(ws, f"B{r}", "NOI", bold=True)
    for i in range(n_years):
        write_input(ws, (r, 3 + i), pf.years[i].noi, fmt="dollar")
    r += 2

    write_section(ws, r, "CapEx"); r += 2
    row("TI", lambda y: -y.ti)
    row("LC", lambda y: -y.lc)
    row("Building CapEx + Reserves", lambda y: -y.building_capex)
    write_label(ws, f"B{r}", "Total CapEx", bold=True)
    for i in range(n_years):
        col_letter = ws.cell(row=r, column=3 + i).column_letter
        write_subtotal(ws, (r, 3 + i), f"=SUM({col_letter}{r-3}:{col_letter}{r-1})", fmt="dollar")
    r += 2

    write_label(ws, f"B{r}", "NCF Unlevered", bold=True)
    for i in range(n_years):
        write_input(ws, (r, 3 + i), pf.years[i].ncf_unlevered, fmt="dollar")
    r += 1
    row("Debt Service", lambda y: -y.debt_service)
    write_label(ws, f"B{r}", "NCF Levered", bold=True)
    for i in range(n_years):
        write_input(ws, (r, 3 + i), pf.years[i].ncf_levered, fmt="dollar")
    r += 2

    write_label(ws, f"B{r}", "Avg Occupancy")
    for i in range(n_years):
        write_input(ws, (r, 3 + i), pf.years[i].avg_occupancy_pct, fmt="pct1")


# ---------------------------------------------------------------------------
# Rollover Schedule
# ---------------------------------------------------------------------------

def _write_rollover(wb: Workbook, pf: CommercialProForma):
    ws = wb.create_sheet("Rollover")
    set_sheet_defaults(ws, "Lease Rollover Schedule")
    write_header(ws, 1, f"{pf.deal.deal_name} - Rollover Schedule", span_cols=8)

    headers = ["Year", "SF Rolling", "% of RBA",
               "In-Place Rent ($)", "Market Rent at Roll ($)", "MTM Spread %"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=3, column=i, value=h).style = "label_bold"

    rba = pf.deal.property.total_rba
    r = 4
    for ro in pf.rollover_schedule:
        ws.cell(row=r, column=2, value=f"Yr {ro.year}")
        write_input(ws, (r, 3), ro.sf_rolling, fmt="general")
        write_input(ws, (r, 4), ro.sf_rolling / rba if rba else 0, fmt="pct1")
        write_input(ws, (r, 5), ro.in_place_rent_rolling, fmt="dollar")
        write_input(ws, (r, 6), ro.market_rent_at_roll, fmt="dollar")
        write_input(ws, (r, 7), ro.mtm_spread_pct, fmt="pct1")
        r += 1

    # Totals
    write_label(ws, f"B{r}", "TOTAL", bold=True)
    total_sf = sum(ro.sf_rolling for ro in pf.rollover_schedule)
    total_inplace = sum(ro.in_place_rent_rolling for ro in pf.rollover_schedule)
    total_market = sum(ro.market_rent_at_roll for ro in pf.rollover_schedule)
    write_input(ws, (r, 3), total_sf, fmt="general")
    write_subtotal(ws, (r, 4), total_sf / rba if rba else 0, fmt="pct1")
    write_subtotal(ws, (r, 5), total_inplace, fmt="dollar")
    write_subtotal(ws, (r, 6), total_market, fmt="dollar")
    if total_inplace > 0:
        write_input(ws, (r, 7), (total_market - total_inplace) / total_inplace, fmt="pct1")


# ---------------------------------------------------------------------------
# Debt + Returns
# ---------------------------------------------------------------------------

def _write_debt(wb: Workbook, pf: CommercialProForma):
    ws = wb.create_sheet("Debt")
    set_sheet_defaults(ws, "Debt Sizing")
    write_header(ws, 1, f"{pf.deal.deal_name} — Debt Sizing & Amort", span_cols=8)

    r = 3
    write_section(ws, r, "Sizing (Yr 1 NOI)"); r += 2
    s = pf.sizing
    for label, val, fmt in [
        ("Loan Amount", s.loan_amount, "dollar"),
        ("Binding Constraint", s.binding, "general"),
        ("Max LTV Capacity", s.constraints.get("LTV", 0), "dollar"),
        ("DSCR Capacity", s.constraints.get("DSCR", 0), "dollar"),
        ("Debt Yield Capacity", s.constraints.get("Debt Yield", 0), "dollar"),
        ("Implied LTV", s.implied_ltv, "pct1"),
        ("Implied DSCR", s.implied_dscr, "multiple"),
        ("Implied Debt Yield", s.implied_debt_yield, "pct2"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Amortization Schedule"); r += 2
    headers = ["Yr", "Beg. Balance", "Interest", "Principal", "Debt Service", "End Balance", "IO?"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=r, column=i, value=h).style = "label_bold"
    r += 1
    for am in pf.amort_schedule:
        ws.cell(row=r, column=2, value=am.year)
        write_input(ws, (r, 3), am.beginning_balance, fmt="dollar")
        write_input(ws, (r, 4), am.interest, fmt="dollar")
        write_input(ws, (r, 5), am.principal, fmt="dollar")
        write_input(ws, (r, 6), am.debt_service, fmt="dollar")
        write_input(ws, (r, 7), am.ending_balance, fmt="dollar")
        ws.cell(row=r, column=8, value="Yes" if am.is_io else "No")
        r += 1


def _write_returns(wb: Workbook, pf: CommercialProForma, wf: WaterfallResult):
    ws = wb.create_sheet("Returns")
    set_sheet_defaults(ws, "Returns")
    write_header(ws, 1, f"{pf.deal.deal_name} — Returns", span_cols=8)

    r = 3
    write_section(ws, r, "Sources & Uses"); r += 2
    su = pf.sources_uses
    for label, val, fmt in [
        ("Purchase Price", su.purchase_price, "dollar"),
        ("Closing Costs", su.closing_costs, "dollar"),
        ("Initial CapEx", su.initial_capex, "dollar"),
        ("Day-One Reserves", su.day_one_reserves, "dollar"),
        ("Acq Fee", su.acq_fee, "dollar"),
        ("Origination Fee", su.origination_fee, "dollar"),
        ("Lender Reserves", su.lender_reserves, "dollar"),
        ("Total Uses", su.total_uses, "dollar"),
        ("Loan Amount", su.loan_amount, "dollar"),
        ("Equity Check", su.equity_check, "dollar"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Exit Summary"); r += 2
    ex = pf.exit_summary
    for label, val, fmt in [
        ("Exit Year", ex.exit_year, "general"),
        ("Exit NOI", ex.exit_noi, "dollar"),
        ("Exit Cap", ex.exit_cap, "pct2"),
        ("Gross Sale Price", ex.gross_sale, "dollar"),
        ("Cost of Sale", ex.cost_of_sale, "dollar"),
        ("Loan Payoff", ex.loan_payoff, "dollar"),
        ("Net Proceeds to Equity", ex.net_proceeds, "dollar"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Returns"); r += 2
    for label, val, fmt in [
        ("Going-In Cap", pf.going_in_cap, "pct2"),
        (f"Yr {pf.roc.stab_yr} Cap (on price)", pf.stabilized_cap, "pct2"),
        (f"Yr {pf.roc.stab_yr} YoC (all-in basis)", pf.roc.trended_stab, "pct2"),
        ("Total Equity IRR (project)", wf.total_equity_irr, "pct2"),
        ("Total Equity MOIC", wf.total_equity_moic, "multiple"),
        ("LP Net IRR", wf.lp.irr, "pct2"),
        ("LP Net MOIC", wf.lp.moic, "multiple"),
        ("GP Net IRR", wf.gp.irr, "pct2"),
        ("GP Net MOIC", wf.gp.moic, "multiple"),
    ]:
        write_label(ws, f"B{r}", label); write_input(ws, f"D{r}", val, fmt=fmt); r += 1

    r += 1; write_section(ws, r, "Equity Cash Flows"); r += 2
    write_label(ws, f"B{r}", "Date", bold=True)
    write_label(ws, f"D{r}", "Amount", bold=True); r += 1
    for ef in pf.equity_flows_total:
        write_input(ws, (r, 2), ef.period, fmt="date")
        write_input(ws, (r, 4), ef.amount, fmt="dollar")
        r += 1


# ---------------------------------------------------------------------------
# Top-level
# ---------------------------------------------------------------------------

def write_commercial_workbook(pf: CommercialProForma, wf: WaterfallResult, out_path: Path) -> Path:
    """Write the full commercial workbook and return the path."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    apply_institutional_styles(wb)

    _write_assumptions(wb, pf)
    _write_rent_roll(wb, pf)
    _write_per_lease(wb, pf)
    _write_pro_forma(wb, pf)
    _write_rollover(wb, pf)
    _write_debt(wb, pf)
    _write_returns(wb, pf, wf)

    from ..excel_summary import build_payload, write_executive_summary
    payload = build_payload(
        pf=pf, wf=wf,
        asset_class=pf.deal.property.asset_class.title(),
        denom_label="RBA (SF)",
        denom_value=pf.deal.property.total_rba,
        per_denom_label="$/SF",
        per_denom_fmt="per_sf",
        value_add_capex_total=0.0,
    )
    write_executive_summary(wb, payload)

    wb.save(out_path)
    return out_path
